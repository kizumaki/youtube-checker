import datetime
import io
import os
import re
import urllib.parse
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter

import isodate
import openpyxl
import pandas as pd
import pycountry
import requests
import streamlit as st
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage

DEFAULT_API_KEY = st.secrets.get("YOUTUBE_API_KEY", "AIzaSyDDBEJscqkGGpG1xtuL4wYPuFkS4BIL854")

ROW_COLORS = [
    "2F5597", "C00000", "70AD47", "7030A0", "00C0C0",
    "E37222", "41536B", "A04000", "385723", "626262"
]

def to_pure_id(raw_val):
    if not raw_val or pd.isna(raw_val) or str(raw_val).strip().upper() in ["N/A", "NAN", "NONE", ""]: return None
    s = str(raw_val).strip()
    
    m_chan = re.search(r'youtube\.com/channel/([a-zA-Z0-9_-]{24})', s)
    if m_chan: return m_chan.group(1)
        
    if re.match(r'^UC[a-zA-Z0-9_-]{22}$', s): return s

    m_url = re.search(r'youtube\.com/(?:@|c/|user/)?([^\s?#/]+)', s, re.IGNORECASE)
    if m_url:
        val = m_url.group(1)
        if val.lower() in ['watch', 'shorts', 'feed', 'embed', 'results']: return None
        s = val
    s = re.sub(r'^@+', '', s).strip().lower()

    pattern = r'_(?:backlog|january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec|\d{1,4})[_\-\s,.]?.*$'
    s = re.sub(pattern, '', s, flags=re.IGNORECASE)

    return s if s else None

def get_channel_link(pure_id):
    if not pure_id: return ""
    if pure_id.startswith('UC') and len(pure_id) == 24:
        return f"https://www.youtube.com/channel/{pure_id}"
    return f"https://www.youtube.com/@{pure_id}"

def extract_search_query(raw_url):
    if not raw_url or pd.isna(raw_url): return None
    s = str(raw_url).strip()
    m = re.search(r'youtube\.com/results\?[^#]*search_query=([^&#]+)', s, re.IGNORECASE)
    if m:
        raw_q = m.group(1)
        decoded = urllib.parse.unquote(raw_q).replace('+', ' ').strip()
        return decoded
    return None

def generate_candidate_handles_from_query(q):
    if not q: return []
    decoded = urllib.parse.unquote(str(q)).strip()
    cleaned_name = re.sub(r'[^\w\s.-]', '', decoded).strip()
    candidates = []
    c1 = re.sub(r'\s+', '', cleaned_name).lower()
    if c1 and len(c1) >= 3: candidates.append(c1)
    c2 = re.sub(r'\s+', '_', cleaned_name).lower()
    c2_clean = re.sub(r'[^a-zA-Z0-9_.-]', '', c2)
    if c2_clean and c2_clean not in candidates and len(c2_clean) >= 3: candidates.append(c2_clean)
    if 'official' in cleaned_name.lower():
        c3 = c1.replace('official', '')
        if c3 and c3 not in candidates and len(c3) >= 3: candidates.append(c3)
    return candidates

def is_garbage_input(s):
    if not s or pd.isna(s): return True
    val = str(s).strip()
    if not val: return True
    if re.match(r'^\d[\d,.]*$', val): return True
    if val.startswith('http') and 'youtube.com' not in val.lower() and 'youtu.be' not in val.lower():
        return True
    return False

def yt_execute_safe(request_func, api_keys, exhausted_keys=None, cost=1):
    if not api_keys: api_keys = [DEFAULT_API_KEY]
    if exhausted_keys is None: exhausted_keys = set()
    
    # Filter out keys that are known to be exhausted
    valid_keys = [k for k in api_keys if k not in exhausted_keys]
    if not valid_keys:
        valid_keys = list(api_keys)
        
    last_err, key_logs = None, []
    for key in valid_keys:
        try:
            yt = build("youtube", "v3", developerKey=key)
            req = request_func(yt)
            res = req.execute()
            key_logs.append((key, "OK", cost))
            return res, key, cost, key_logs
        except HttpError as e:
            status_code = e.resp.status if hasattr(e, 'resp') else 0
            err_content = str(e.content).lower() if hasattr(e, 'content') else str(e).lower()
            
            # Phân biệt chính xác lỗi hết Quota / Key hỏng thực sự (403, 429)
            is_quota_or_key_err = (status_code in [403, 429]) or ("quota" in err_content) or ("key" in err_content and "not valid" in err_content)
            
            if is_quota_or_key_err:
                exhausted_keys.add(key)
                key_logs.append((key, "EXHAUSTED", 0))
                last_err = e
                continue
            else:
                # Lỗi tham số / không tìm thấy (400, 404): Giữ nguyên Key sống, trả kết quả rỗng
                return {"items": []}, key, 0, key_logs
        except Exception as e:
            last_err = e
            continue
            
    if last_err: raise last_err
    return {"items": []}, valid_keys[0] if valid_keys else DEFAULT_API_KEY, 0, key_logs

def test_all_api_keys():
    keys = st.session_state.get('api_keys', [])
    usage = st.session_state.get('api_usage', {})
    status_map = st.session_state.get('api_status_map', {})
    exhausted_set = st.session_state.get('exhausted_keys_set', set())
    for k in keys:
        try:
            yt = build("youtube", "v3", developerKey=k)
            yt.channels().list(part="id", id="UC_x5XG1OV2P6uZZ5FSM9Ttw").execute()
            status_map[k] = ("OK", usage.get(k, 0))
            if k in exhausted_set: exhausted_set.remove(k)
        except Exception:
            usage[k] = 10000
            status_map[k] = ("DEAD", 10000)
            exhausted_set.add(k)
    st.session_state['api_usage'] = usage
    st.session_state['api_status_map'] = status_map
    st.session_state['exhausted_keys_set'] = exhausted_set

def get_channel_id_by_handle_direct(handle, api_keys, exhausted_keys=None):
    clean = handle.replace('@', '').split('/')[-1].strip()
    if re.match(r'^UC[a-zA-Z0-9_-]{22}$', clean):
        try:
            res, key_used, cost, logs = yt_execute_safe(lambda yt: yt.channels().list(part="id", id=clean), api_keys, exhausted_keys, cost=1)
            if 'items' in res and len(res['items']) > 0:
                return res['items'][0]['id'], key_used, cost, logs
        except Exception: pass
    try:
        res, key_used, cost, logs = yt_execute_safe(lambda yt: yt.channels().list(part="id", forHandle=clean.lower()), api_keys, exhausted_keys, cost=1)
        if 'items' in res and len(res['items']) > 0:
            return res['items'][0]['id'], key_used, cost, logs
    except Exception: pass
    try:
        res, key_used, cost, logs = yt_execute_safe(lambda yt: yt.search().list(part="snippet", q=clean, type="channel", maxResults=1), api_keys, exhausted_keys, cost=100)
        if 'items' in res and len(res['items']) > 0:
            return res['items'][0]['snippet']['channelId'], key_used, cost, logs
    except Exception: pass
    return None, None, 0, []

def get_channel_details_direct(channel_id, api_keys, exhausted_keys=None):
    res, key_used, cost, logs = yt_execute_safe(lambda yt: yt.channels().list(part="snippet,contentDetails,statistics", id=channel_id), api_keys, exhausted_keys, cost=1)
    if 'items' in res and len(res['items']) > 0:
        item = res['items'][0]
        playlist_id = item.get('contentDetails', {}).get('relatedPlaylists', {}).get('uploads', '')
        sub_count = int(item['statistics'].get('subscriberCount', 0))
        description = item['snippet'].get('description', 'No description available.')
        joined_date_raw = item['snippet'].get('publishedAt', '')
        joined_date = pd.to_datetime(joined_date_raw).strftime("%b %d, %Y") if joined_date_raw else ""
        country_code = item['snippet'].get('country', 'N/A')
        country_name = pycountry.countries.get(alpha_2=country_code).name if country_code != 'N/A' and pycountry.countries.get(alpha_2=country_code) else country_code
        avatar_url = item['snippet'].get('thumbnails', {}).get('high', {}).get('url', '')
        channel_title = item['snippet'].get('title', '')
        custom_url = item['snippet'].get('customUrl', '')
        return playlist_id, sub_count, description, joined_date, country_name, country_code, avatar_url, channel_title, custom_url, key_used, cost, logs
    return None, 0, "", "", "", "", "", "", "", None, 0, []

def get_video_details_direct(video_ids, api_keys, exhausted_keys=None):
    video_data, logs = [], []
    for i in range(0, len(video_ids), 50):
        try:
            chunk = video_ids[i:i+50]
            res, key_used, cost, l_chunk = yt_execute_safe(lambda yt: yt.videos().list(part="snippet,contentDetails,statistics", id=','.join(chunk)), api_keys, exhausted_keys, cost=1)
            logs.extend(l_chunk)
            for item in res.get('items', []):
                duration_seconds = int(isodate.parse_duration(item['contentDetails']['duration']).total_seconds())
                h, rem = divmod(duration_seconds, 3600)
                m, s = divmod(rem, 60)
                dur_str = f"{h:02d}:{m:02d}:{s:02d}"
                pub_date = item['snippet']['publishedAt']
                formatted_date = pd.to_datetime(pub_date).strftime("%Y-%m-%d")
                video_data.append({
                    'Title': item['snippet']['title'], 'Description': item['snippet'].get('description', ''),
                    'Link': f"https://youtube.com/watch?v={item['id']}", 'Length (Exact)': dur_str, 
                    'Seconds': duration_seconds, 'Views': int(item['statistics'].get('viewCount', 0)), 
                    'Published Date': formatted_date, 'Video ID': item['id']
                })
        except Exception: pass
    return video_data, logs

def is_long_form_video(v, min_seconds=180):
    title = v.get('Title', '').lower()
    if '#shorts' in title or '#short' in title: return False
    return v.get('Seconds', 0) > min_seconds

def get_6_recent_videos_direct(pure_handle, cid, api_keys, exhausted_keys=None):
    long_vids = []
    try:
        if cid:
            playlist_id, _, _, _, _, _, _, _, _, _ = get_channel_details_direct(cid, api_keys, exhausted_keys)
            if playlist_id:
                v_res, _, _, _ = yt_execute_safe(lambda yt: yt.playlistItems().list(part="snippet", playlistId=playlist_id, maxResults=50), api_keys, exhausted_keys, cost=1)
                v_ids = [v_item['snippet']['resourceId']['videoId'] for v_item in v_res.get('items', [])]
                if v_ids:
                    v_details, _ = get_video_details_direct(v_ids, api_keys, exhausted_keys)
                    for v in v_details:
                        if is_long_form_video(v, min_seconds=180): long_vids.append(v)
                        if len(long_vids) >= 6: break
    except Exception: pass
    return long_vids[:6]

def extract_contacts_and_socials(text_corpus):
    if not text_corpus: return {}
    corpus = str(text_corpus)
    contacts = {}
    emails = re.findall(r'\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b', corpus)
    valid_emails = [e for e in emails if not any(e.lower().endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg']) and e.lower() not in ['user@domain.com', 'info@youtube.com']]
    if valid_emails: contacts['Email'] = valid_emails[0]
    
    ig = re.findall(r'(?:https?://)?(?:www\.)?instagram\.com/([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if ig: contacts['Instagram'] = f"https://instagram.com/{ig[0].rstrip('./-_,')}"
    tt = re.findall(r'(?:https?://)?(?:www\.)?tiktok\.com/@?([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if tt: contacts['TikTok'] = f"https://tiktok.com/@{tt[0].rstrip('./-_,')}"
    x = re.findall(r'(?:https?://)?(?:www\.)?(?:twitter\.com|x\.com)/([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if x: contacts['Twitter'] = f"https://x.com/{x[0].rstrip('./-_,')}"
    dc = re.findall(r'(?:https?://)?(?:www\.)?discord\.(?:gg|com/invite)/([a-zA-Z0-9_-]+)', corpus, re.IGNORECASE)
    if dc: contacts['Discord'] = f"https://discord.gg/{dc[0]}"
    fb = re.findall(r'(?:https?://)?(?:www\.)?facebook\.com/([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if fb: contacts['Facebook'] = f"https://facebook.com/{fb[0].rstrip('./-_,')}"
    return contacts

def render_social_badges_html(contacts_dict):
    if not contacts_dict: return "<span style='font-size:0.75rem; color:#9CA3AF;'>Chưa có thông tin MXH</span>"
    html = "<div style='margin-top: 6px;'>"
    if 'Email' in contacts_dict: html += f"<a href='mailto:{contacts_dict['Email']}' class='social-badge social-email'>✉️ {contacts_dict['Email']}</a>"
    if 'Instagram' in contacts_dict: html += f"<a href='{contacts_dict['Instagram']}' target='_blank' class='social-badge social-ig'>📸 IG</a>"
    if 'TikTok' in contacts_dict: html += f"<a href='{contacts_dict['TikTok']}' target='_blank' class='social-badge social-tt'>🎵 TikTok</a>"
    if 'Twitter' in contacts_dict: html += f"<a href='{contacts_dict['Twitter']}' target='_blank' class='social-badge social-x'>🐦 X</a>"
    if 'Discord' in contacts_dict: html += f"<a href='{contacts_dict['Discord']}' target='_blank' class='social-badge social-discord'>💬 Discord</a>"
    if 'Facebook' in contacts_dict: html += f"<a href='{contacts_dict['Facebook']}' target='_blank' class='social-badge social-fb'>📘 FB</a>"
    if 'Website' in contacts_dict: html += f"<a href='{contacts_dict['Website']}' target='_blank' class='social-badge social-web'>🌐 Web</a>"
    html += "</div>"
    return html

def process_single_crm_channel_meta(pure_handle, api_keys, exhausted_keys=None):
    if not pure_handle: return pure_handle, {"sub_count": -1, "sub_str": "N/A", "country": "N/A", "socials": {}}, []
    logs = []
    try:
        cid, k1, c1, l1 = get_channel_id_by_handle_direct(pure_handle, api_keys, exhausted_keys)
        logs.extend(l1)
        if cid:
            playlist_id, sub_count, desc, joined, country_name, country_code, avatar, k2, c2, l2 = get_channel_details_direct(cid, api_keys, exhausted_keys)
            logs.extend(l2)
            recent_vids = get_6_recent_videos_direct(pure_handle, cid, api_keys, exhausted_keys)
            v_descs = " ".join([v.get('Description', '') for v in recent_vids]) if recent_vids else ""
            socials = extract_contacts_and_socials(f"{desc} {v_descs}")
            return pure_handle, {"sub_count": sub_count, "sub_str": f"{sub_count:,}" if sub_count > 0 else "N/A", "country": country_name if country_name else "N/A", "socials": socials}, logs
    except Exception: pass
    return pure_handle, {"sub_count": -1, "sub_str": "N/A", "country": "N/A", "socials": {}}, logs

def extract_video_id(raw_url):
    if not raw_url or pd.isna(raw_url): return None
    m = re.search(r'(?:v=|\/shorts\/|youtu\.be\/)([a-zA-Z0-9_-]{11})', str(raw_url).strip())
    return m.group(1) if m else None

def get_handles_from_video_ids(video_ids):
    if not video_ids: return []
    active_keys = st.session_state.get('api_keys', [DEFAULT_API_KEY])
    channel_ids = set()
    for i in range(0, len(video_ids), 50):
        try:
            res, _, _, _ = yt_execute_safe(lambda yt: yt.videos().list(part="snippet", id=','.join(video_ids[i:i+50])), active_keys, cost=1)
            for item in res.get('items', []):
                c_id = item.get('snippet', {}).get('channelId')
                if c_id: channel_ids.add(c_id)
        except Exception: pass
        
    handles = []
    if channel_ids:
        c_list = list(channel_ids)
        for i in range(0, len(c_list), 50):
            try:
                res, _, _, _ = yt_execute_safe(lambda yt: yt.channels().list(part="snippet", id=','.join(c_list[i:i+50])), active_keys, cost=1)
                for item in res.get('items', []):
                    custom_url = item.get('snippet', {}).get('customUrl', '')
                    pure = to_pure_id(custom_url) or to_pure_id(item.get('id'))
                    if pure and pure not in handles: handles.append(pure)
            except Exception: pass
    return handles

def get_handles_from_search_queries(search_queries):
    if not search_queries: return []
    active_keys = st.session_state.get('api_keys', [DEFAULT_API_KEY])
    exhausted_set = set(st.session_state.get('exhausted_keys_set', set()))
    handles = []
    for q in search_queries:
        if not q or not str(q).strip(): continue
        decoded_q = urllib.parse.unquote(str(q)).strip()
        try:
            res, _, _, _ = yt_execute_safe(lambda yt: yt.search().list(part="snippet", q=decoded_q, type="channel", maxResults=1), active_keys, exhausted_set, cost=100)
            if res.get('items'):
                c_id = res['items'][0]['snippet']['channelId']
                c_res, _, _, _ = yt_execute_safe(lambda yt: yt.channels().list(part="snippet", id=c_id), active_keys, exhausted_set, cost=1)
                if c_res.get('items'):
                    custom_url = c_res['items'][0]['snippet'].get('customUrl', '')
                    pure = to_pure_id(custom_url) or to_pure_id(c_id)
                    if pure and pure not in handles: handles.append(pure)
        except Exception: pass
    return handles

def parse_raw_inputs_to_handles(raw_inputs_list):
    handles, video_ids, search_queries = set(), set(), set()
    for raw in raw_inputs_list:
        if is_garbage_input(raw): continue
        s = str(raw).strip()
        if not s: continue
        sq = extract_search_query(s)
        if sq:
            search_queries.add(sq)
            for cand in generate_candidate_handles_from_query(sq): handles.add(cand)
            continue
        v_id = extract_video_id(s)
        if v_id: video_ids.add(v_id); continue
        
        uc_cids = re.findall(r'\b(UC[a-zA-Z0-9_-]{22})\b', s)
        if uc_cids:
            for cid in uc_cids: handles.add(cid)
            continue
            
        at_handles = re.findall(r'@([a-zA-Z0-9_.-]{3,})', s)
        if at_handles:
            for h in at_handles: handles.add(h.lower())
            continue

        if s.startswith('@') or 'youtube.com/' in s.lower() or re.match(r'^UC[a-zA-Z0-9_-]{22}$', s):
            p_h = to_pure_id(s)
            if p_h: handles.add(p_h)
            continue
            
        clean_text = re.sub(r'^@+', '', s).strip()
        if clean_text:
            search_queries.add(clean_text)
            for cand in generate_candidate_handles_from_query(clean_text): handles.add(cand)

    if search_queries:
        try:
            with st.spinner(f"🔍 Đang truy vấn YouTube Search API cho {len(search_queries)} tên YouTuber..."):
                for h in get_handles_from_search_queries(list(search_queries)): handles.add(h)
        except Exception: pass

    if video_ids:
        try:
            with st.spinner(f"🔍 Đang giải mã {len(video_ids)} Link Video sang Handle..."):
                for h in get_handles_from_video_ids(list(video_ids)): handles.add(h)
        except Exception: pass
            
    return list(handles)

def extract_text_from_docx_bytes(file_bytes):
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            xml_content = z.read('word/document.xml')
            tree = ET.fromstring(xml_content)
            return " ".join([elem.text for elem in tree.iter() if elem.tag.endswith('}t') and elem.text])
    except Exception: return ""

def extract_raw_inputs_from_file(uploaded_file):
    raw_list, fname = [], uploaded_file.name.lower()
    try:
        uploaded_file.seek(0); file_bytes = uploaded_file.read(); uploaded_file.seek(0)
        if fname.endswith('.txt'): raw_list = re.split(r'[\n,\t\r]+', file_bytes.decode("utf-8", errors="ignore"))
        elif fname.endswith('.csv') or fname.endswith('.xls') or fname.endswith('.xlsx'):
            df = pd.read_csv(io.BytesIO(file_bytes)) if fname.endswith('.csv') else pd.read_excel(io.BytesIO(file_bytes))
            target_cols = [col for col in df.columns if any(k in str(col).lower() for k in ['search', 'youtuber', 'handle', 'link', 'kênh'])]
            cols_to_use = target_cols if target_cols else df.columns
            for col in cols_to_use:
                if 'stats' in str(col).lower() or 'kz.youtubers' in str(col).lower(): continue
                for val in df[col].dropna(): raw_list.append(str(val))
        elif fname.endswith('.docx') or fname.endswith('.doc'):
            raw_list = re.split(r'[\n,\t\r]+', extract_text_from_docx_bytes(file_bytes))
    except Exception as e: st.error(f"Lỗi đọc file: {e}")
    return raw_list

def extract_handle_from_filename(filename):
    base = os.path.basename(filename)
    base_no_ext = os.path.splitext(base)[0]
    pattern = r'_(?:backlog|january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec|\d{1,4})[_\-\s,.]?.*$'
    cleaned = re.sub(pattern, '', base_no_ext, flags=re.IGNORECASE)
    pure_id = re.sub(r'^@+', '', cleaned).strip().lower()
    return pure_id if pure_id else None

def is_within_last_90_days(date_str):
    if not date_str or date_str == "N/A": return False
    s, today = str(date_str).strip(), datetime.date.today()
    m_dmy = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})', s)
    if m_dmy:
        try: return 0 <= (today - datetime.date(int(m_dmy.group(3)), int(m_dmy.group(2)), int(m_dmy.group(1)))).days <= 90
        except Exception: pass
    m_ymd = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
    if m_ymd:
        try: return 0 <= (today - datetime.date(int(m_ymd.group(1)), int(m_ymd.group(2)), int(m_ymd.group(3)))).days <= 90
        except Exception: pass
    try: return 0 <= (today - pd.to_datetime(s, dayfirst=True).date()).days <= 90
    except Exception: return False

EXCLUDED_COUNTRIES = {'CN', 'TW', 'HK', 'TH', 'IN', 'VN'}
NON_LATIN_REGEX = re.compile(r'[฀-๿]|[一-鿿]|[ऀ-ॿ]|[가-힯]', re.IGNORECASE)
VIETNAMESE_UNIQUE_REGEX = re.compile(r'[ơờớởỡợưừứửữựđĐăằắẳẵặảẻỉỏủỷạẹịọụỵềếểễệồốổỗộầấẩẫậ]', re.IGNORECASE)
EXCLUDED_KEYWORDS = ['official mv', 'music video', 'official audio', 'album', 'song', 'records', 'lyrics', 'remix', 'news', 'tin tức', 'lgbt', 'lgbtq', 'war']

def passes_layer1_metadata_filter(title, desc, country_code):
    if country_code in EXCLUDED_COUNTRIES: return False, f"Quốc gia bị loại ({country_code})"
    text = f"{title} {desc}".lower()
    if NON_LATIN_REGEX.search(text): return False, "Ngôn ngữ không phù hợp"
    if VIETNAMESE_UNIQUE_REGEX.search(text): return False, "Kênh Ngôn Ngữ Tiếng Việt"
    for kw in EXCLUDED_KEYWORDS:
        if re.search(r'\b' + re.escape(kw) + r'\b', text): return False, f"Loại nội dung cấm ({kw.upper()})"
    return True, "OK"

def clean_and_extract_keywords(text, seed_handle=""):
    if not text or not str(text).strip():
        return []
    
    seed_clean = seed_handle.replace('@', '').strip().lower()
    
    if ',' in str(text):
        raw_terms = [t.strip().lower() for t in str(text).split(',') if t.strip()]
    else:
        raw_terms = [t.strip().lower() for t in re.split(r'[\n\r\t]+', str(text)) if t.strip()]

    clean_terms = []
    stop_words = {'the', 'a', 'and', 'official', 'channel', 'video', 'youtube', 'shorts', 'short'}
    
    for term in raw_terms:
        t = term.strip()
        if not t or len(t) < 2:
            continue
        if t in stop_words or t == seed_clean:
            continue
        if t not in clean_terms:
            clean_terms.append(t)
            
    return clean_terms

def extract_channel_master_keywords(channel_id):
    active_keys = st.session_state.get('api_keys', [DEFAULT_API_KEY])
    pool, channel_kw, video_tags = [], [], []
    try:
        ch_res, _, _, _ = yt_execute_safe(lambda yt: yt.channels().list(part="brandingSettings,contentDetails", id=channel_id), active_keys, cost=1)
        if 'items' in ch_res and len(ch_res['items']) > 0:
            item = ch_res['items'][0]
            raw_kw = item.get('brandingSettings', {}).get('channel', {}).get('keywords', '')
            found = re.findall(r'"([^"]+)"|\b([a-zA-Z0-9]{3,})\b', raw_kw)
            for k1, k2 in found:
                kw = k1 or k2
                if kw and len(kw) > 2:
                    pool.append(kw.lower())
                    channel_kw.append(kw.lower())
            
            playlist_id = item.get('contentDetails', {}).get('relatedPlaylists', {}).get('uploads', '')
            if playlist_id:
                v_res, _, _, _ = yt_execute_safe(lambda yt: yt.playlistItems().list(part="snippet", playlistId=playlist_id, maxResults=15), active_keys, cost=1)
                v_ids = [v_item['snippet']['resourceId']['videoId'] for v_item in v_res.get('items', []) if v_item.get('snippet', {}).get('resourceId', {}).get('videoId')]
                if v_ids:
                    v_details, _, _, _ = yt_execute_safe(lambda yt: yt.videos().list(part="snippet", id=','.join(v_ids[:15])), active_keys, cost=1)
                    tag_counter = Counter()
                    for v_item in v_details.get('items', []):
                        tags = v_item.get('snippet', {}).get('tags', [])
                        for t in tags:
                            t_clean = t.strip().lower()
                            if len(t_clean) > 2:
                                tag_counter[t_clean] += 1
                                pool.append(t_clean)
                    video_tags = [t for t, _ in tag_counter.most_common(10)]
    except Exception: pass
    
    master = [w for w, _ in Counter(pool).most_common(15)]
    return {
        "master_keywords": master,
        "channel_keywords": list(dict.fromkeys(channel_kw))[:10],
        "top_tags": video_tags
    }

def generate_v414_excel_report(clean_handle, sub_count, channel_desc, channel_joined, channel_country, avatar_url, video_data):
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: Channel Summary Report
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = clean_handle.upper()[:31]
    
    # Header Banner
    ws1.merge_cells('A1:E1')
    report_title = f"{clean_handle.upper()} YOUTUBE CHANNEL SUMMARY REPORT - up to {datetime.datetime.now().strftime('%d-%m-%Y')}"
    ws1['A1'] = report_title
    ws1['A1'].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws1['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Calculate metrics
    total_videos = len(video_data)
    total_views = sum(v.get('Views', 0) for v in video_data)
    total_seconds = sum(v.get('Seconds', 0) for v in video_data)
    total_minutes = round(total_seconds / 60)
    
    # Summary Metrics (Rows 2 to 7)
    ws1['A2'] = f"Total Videos: {total_videos:,}"
    ws1['A2'].font = Font(bold=True)
    
    ws1['A3'] = f"Total Duration: {total_minutes:,} minutes"
    ws1['A3'].font = Font(bold=True)
    
    ws1['A4'] = f"Total Views: {total_views:,}"
    ws1['A4'].font = Font(bold=True)
    
    ws1['A5'] = f"Total Subscribers: {sub_count:,}"
    ws1['A5'].font = Font(bold=True)
    
    ws1['A6'] = f"Country Location: {channel_country if channel_country else 'N/A'}"
    ws1['A6'].font = Font(bold=True)
    
    ws1['A7'] = f"Channel Joined Date: {channel_joined if channel_joined else 'N/A'}"
    ws1['A7'].font = Font(bold=True)
    
    # Description Block
    ws1['A9'] = "Channel Description Text:"
    ws1['A9'].font = Font(bold=True)
    
    ws1['A10'] = channel_desc if channel_desc else "N/A"
    ws1['A10'].alignment = Alignment(wrap_text=True, vertical="top")
    
    # Embed Avatar Image at C2
    if avatar_url:
        try:
            img_res = requests.get(avatar_url, timeout=5)
            if img_res.status_code == 200:
                img_data = io.BytesIO(img_res.content)
                pil_img = PILImage.open(img_data)
                pil_img = pil_img.resize((140, 140))
                
                img_buf = io.BytesIO()
                pil_img.save(img_buf, format="PNG")
                img_buf.seek(0)
                
                from openpyxl.drawing.image import Image as ExcelImage
                excel_img = ExcelImage(img_buf)
                ws1.add_image(excel_img, 'C2')
        except Exception: pass

    # Row 12: Video Table Headers
    headers = ["Video Title", "Link", "Length", "Views", "Published Date"]
    for c_idx, h_text in enumerate(headers, 1):
        cell = ws1.cell(row=12, column=c_idx, value=h_text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[12].height = 20

    # Row 13+: Video Table Data
    for idx, v in enumerate(video_data):
        r = idx + 13
        c_title = ws1.cell(row=r, column=1, value=v.get('Title', ''))
        c_link = ws1.cell(row=r, column=2, value=v.get('Link', ''))
        
        # Make Column B Clickable Hyperlink
        if v.get('Link'):
            c_link.hyperlink = v.get('Link')
            c_link.font = Font(color="0000FF", underline="single")
        
        dur_sec = v.get('Seconds', 0)
        h, rem = divmod(dur_sec, 3600)
        m, s = divmod(rem, 60)
        dur_str = f"{h:02d}:{m:02d}:{s:02d}"
        
        c_len = ws1.cell(row=r, column=3, value=dur_str)
        c_views = ws1.cell(row=r, column=4, value=int(v.get('Views', 0)))
        c_views.number_format = '#,##0'
        c_date = ws1.cell(row=r, column=5, value=str(v.get('Published Date', '')))
        
        c_len.alignment = Alignment(horizontal="center", vertical="center")

    ws1.column_dimensions['A'].width = 55
    ws1.column_dimensions['B'].width = 45
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15

    # -------------------------------------------------------------
    # SHEET 2: Top 10 Video Title
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Top 10 Video Title")
    
    c1 = ws2.cell(row=1, column=1, value="Top 10 Most Viewed Videos (Click to Watch)")
    c1.font = Font(bold=True, color="FFFFFF")
    c1.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    c2 = ws2.cell(row=1, column=2, value="Views")
    c2.font = Font(bold=True, color="FFFFFF")
    c2.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    ws2.cell(row=1, column=4, value=f"📊 Top 10 Most Viewed Videos - {clean_handle.upper()}").font = Font(size=14, bold=True, color="1F4E78")
    
    top_10 = sorted(video_data, key=lambda x: x.get('Views', 0), reverse=True)[:10]
    
    for idx, v in enumerate(top_10):
        r = idx + 2
        color = ROW_COLORS[idx % len(ROW_COLORS)]
        
        cell_title = ws2.cell(row=r, column=1, value=v.get('Title', ''))
        cell_title.font = Font(bold=True, color="FFFFFF")
        cell_title.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        if v.get('Link'):
            cell_title.hyperlink = v.get('Link')
            
        cell_views = ws2.cell(row=r, column=2, value=int(v.get('Views', 0)))
        cell_views.font = Font(bold=True, color="FFFFFF")
        cell_views.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_views.number_format = '#,##0'
        cell_views.alignment = Alignment(horizontal="right")

    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 15

    # Bar Chart on Sheet 2
    if len(top_10) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.y_axis.title = "Total Views"
        chart.x_axis.title = "Videos"
        chart.legend = None
        chart.width = 20
        chart.height = 12
        
        data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(top_10) + 1)
        cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(top_10) + 1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        series = chart.series[0]
        for idx in range(len(top_10)):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = ROW_COLORS[idx % len(ROW_COLORS)]
            series.dPt.append(dp)
        
        ws2.add_chart(chart, "D3")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def run_single_channel_audit(pure_handle, api_keys, exhausted_keys=None):
    logs = []
    try:
        cid, _, _, l1 = get_channel_id_by_handle_direct(pure_handle, api_keys, exhausted_keys)
        logs.extend(l1)
        if not cid: return None, None, logs
        res_det = get_channel_details_direct(cid, api_keys, exhausted_keys)
        playlist_id, sub_count, desc, joined, country, c_code, avatar, channel_title, custom_url = res_det[:9]
        logs.extend(res_det[-1])
        if not playlist_id: return None, None, logs

        clean_custom = to_pure_id(custom_url) or ""
        display_handle = clean_custom if clean_custom else (channel_title if channel_title else pure_handle)

        v_ids, next_token = [], None
        for _ in range(20):
            res, _, _, l3 = yt_execute_safe(lambda yt: yt.playlistItems().list(part="snippet", playlistId=playlist_id, maxResults=50, pageToken=next_token), api_keys, exhausted_keys, cost=1)
            logs.extend(l3)
            items = res.get('items', [])
            if not items: break
            for v_item in items:
                v_id = v_item.get('snippet', {}).get('resourceId', {}).get('videoId')
                if v_id: v_ids.append(v_id)
            next_token = res.get('nextPageToken')
            if not next_token: break

        if not v_ids: return None, None, logs
        v_data, l4 = get_video_details_direct(v_ids, api_keys, exhausted_keys)
        logs.extend(l4)
        
        excel_bytes = generate_v414_excel_report(display_handle, sub_count, desc, joined, country, avatar, v_data)
        return excel_bytes, f"{display_handle}_{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx", logs
    except Exception: return None, None, logs

def process_tab1_single_handle(p_id, db_existing_map, api_keys, exhausted_keys=None):
    p_clean = p_id.lower()
    
    # 1. Check direct input against DB map
    if p_clean in db_existing_map:
        db_item = db_existing_map[p_clean]
        return "EXISTING", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": db_item.get("youtuber_name", p_id.upper()), "Trạng thái": "❌ Đã có trong DB"}, []
    
    cid, _, _, l1 = get_channel_id_by_handle_direct(p_clean, api_keys, exhausted_keys)
    logs = list(l1)
    if not cid: return "REJECTED", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": p_id.upper(), "Trạng thái": "❌ Không tìm thấy kênh", "Lý do loại": "Không tồn tại trên YT"}, logs
    
    # 2. Check Channel ID against DB map
    cid_clean = cid.lower()
    if cid_clean in db_existing_map:
        db_item = db_existing_map[cid_clean]
        return "EXISTING", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": db_item.get("youtuber_name", p_id.upper()), "Trạng thái": "❌ Đã có trong DB"}, []

    try:
        res, _, _, l2 = yt_execute_safe(lambda yt: yt.channels().list(part="snippet,statistics", id=cid), api_keys, exhausted_keys, cost=1)
        logs.extend(l2)
        if res.get('items'):
            item = res['items'][0]
            custom_url = item['snippet'].get('customUrl', '')
            pure_custom = to_pure_id(custom_url)
            
            # 3. Check Custom URL Handle against DB map
            if pure_custom and pure_custom.lower() in db_existing_map:
                db_item = db_existing_map[pure_custom.lower()]
                return "EXISTING", {"Handle": f"@{pure_custom}", "Tên Kênh": db_item.get("youtuber_name", item['snippet'].get('title', p_id.upper())), "Trạng thái": "❌ Đã có trong DB"}, logs

            ch_title = item['snippet'].get('title', p_id.upper())
            ch_desc = item['snippet'].get('description', '')
            country_code = item['snippet'].get('country', 'N/A')
            country_name = pycountry.countries.get(alpha_2=country_code).name if country_code != 'N/A' and pycountry.countries.get(alpha_2=country_code) else country_code
            sub_count = int(item['statistics'].get('subscriberCount', 0))
            recent_vids = get_6_recent_videos_direct(p_id, cid, api_keys, exhausted_keys)
            socials = extract_contacts_and_socials(f"{ch_title} {ch_desc} " + " ".join([v.get('Description', '') for v in recent_vids]))

            if sub_count < 1000000:
                return "REJECTED", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": ch_title, "Subscribers": f"{sub_count:,}", "Trạng thái": f"❌ Dưới 1 triệu Subs ({sub_count:,})", "Lý do loại": f"Dưới 1M Subs ({sub_count:,})", "Socials": socials}, logs
            
            passes_l1, l1_reason = passes_layer1_metadata_filter(ch_title, ch_desc, country_code)
            if not passes_l1:
                return "REJECTED", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": ch_title, "Subscribers": f"{sub_count:,}", "Trạng thái": f"❌ {l1_reason}", "Lý do loại": l1_reason, "Socials": socials}, logs

            if not recent_vids:
                return "REJECTED", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": ch_title, "Subscribers": f"{sub_count:,}", "Trạng thái": "❌ Kênh chỉ làm Shorts / Không có video dài", "Lý do loại": "Kênh chỉ làm Shorts", "Socials": socials}, logs

            latest_date = recent_vids[0].get('Published Date', 'N/A')
            if not is_within_last_90_days(latest_date):
                return "REJECTED", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": ch_title, "Subscribers": f"{sub_count:,}", "Trạng thái": f"❌ Ngưng hoạt động (>90 ngày, gần nhất: {latest_date})", "Lý do loại": "Kênh ngưng hoạt động (>90 ngày)", "Socials": socials}, logs

            if not any(v.get('Seconds', 0) >= 600 for v in recent_vids):
                return "REJECTED", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": ch_title, "Subscribers": f"{sub_count:,}", "Trạng thái": "❌ Không có video > 10 phút trong các video gần nhất", "Lý do loại": "Không có video > 10 phút", "Socials": socials}, logs
            
            return "NEW", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": ch_title, "Subscribers": f"{sub_count:,}", "Quốc gia": country_name, "Link Kênh": get_channel_link(p_id), "Trạng thái": "✅ Kênh Mới Đạt Chuẩn", "Socials": socials}, logs
    except Exception: pass
    return "REJECTED", {"Handle": f"@{p_id}" if not p_id.startswith('UC') else p_id, "Tên Kênh": p_id.upper(), "Trạng thái": "❌ Lỗi đọc dữ liệu API", "Lý do loại": "Lỗi API", "Socials": {}}, logs

def check_is_existing(c_handle, c_id, c_custom_url, c_title, db_existing_set):
    if not db_existing_set: return False
    candidates_to_check = set()
    if c_handle:
        h_clean = str(c_handle).strip().lower()
        candidates_to_check.add(h_clean)
        candidates_to_check.add(re.sub(r'^@+', '', h_clean))
        p = to_pure_id(c_handle)
        if p:
            candidates_to_check.add(p.lower())
            candidates_to_check.add(f"@{p.lower()}")
            
    if c_id:
        candidates_to_check.add(str(c_id).strip().lower())
        
    if c_custom_url:
        cu_clean = str(c_custom_url).strip().lower()
        candidates_to_check.add(cu_clean)
        candidates_to_check.add(re.sub(r'^@+', '', cu_clean))
        p_cu = to_pure_id(c_custom_url)
        if p_cu:
            candidates_to_check.add(p_cu.lower())
            
    if c_title:
        t_clean = str(c_title).strip().lower()
        candidates_to_check.add(t_clean)
        candidates_to_check.add(re.sub(r'[^a-zA-Z0-9_.-]', '', t_clean))

    for cand in candidates_to_check:
        if cand and cand in db_existing_set:
            return True
    return False

def process_single_candidate(item, min_subs_choice, min_duration_choice, db_existing_set, api_keys, exhausted_keys=None):
    c_handle = to_pure_id(item['snippet'].get('customUrl', '')) or item['id']
    c_title, c_desc, c_country = item['snippet']['title'], item['snippet'].get('description', ''), item['snippet'].get('country', 'N/A')
    c_subs, c_video_count = int(item['statistics'].get('subscriberCount', 0)), int(item['statistics'].get('videoCount', 0))
    c_url = get_channel_link(c_handle)
    db_status = "❌ Đã có trong DB" if check_is_existing(c_handle, item.get('id'), item.get('snippet', {}).get('customUrl'), c_title, db_existing_set) else "✅ KÊNH MỚI"
    
    c_playlist = item.get('contentDetails', {}).get('relatedPlaylists', {}).get('uploads', '')
    latest_date, recent_vids = "N/A", []
    avg_views, er_rate, score = 0, 0, 0
    if c_playlist and c_video_count > 0:
        try:
            v_res, _, _, _ = yt_execute_safe(lambda yt: yt.playlistItems().list(part="snippet", playlistId=c_playlist, maxResults=50), api_keys, exhausted_keys, cost=1)
            v_ids = [v_item['snippet']['resourceId']['videoId'] for v_item in v_res.get('items', [])]
            if v_ids:
                v_details, _ = get_video_details_direct(v_ids, api_keys, exhausted_keys)
                long_vids = [v for v in v_details if is_long_form_video(v, min_seconds=180)]
                if long_vids:
                    latest_date = long_vids[0]['Published Date']
                    recent_vids = long_vids[:6]
                    if recent_vids and c_subs > 0:
                        avg_views = sum(v.get('Views', 0) for v in recent_vids) / len(recent_vids)
                        er_rate = (avg_views / c_subs) * 100
                        score = min(100, int((er_rate / 10.0) * 100))
        except Exception: pass

    socials = extract_contacts_and_socials(f"{c_title} {c_desc} " + " ".join([v.get('Description', '') for v in recent_vids]))
    base_data = {
        "Handle": f"@{c_handle}" if not c_handle.startswith('UC') else c_handle, "Link Kênh": c_url, "Tên Kênh": c_title, 
        "Subscribers": f"{c_subs:,}", "Quốc gia": c_country, "Video Gần Nhất": latest_date, "Tổng Số Video": f"{c_video_count:,}", 
        "Trạng Thái DB": db_status, "recent_videos": recent_vids, "ER": f"{er_rate:.2f}%" if er_rate > 0 else "N/A",
        "Score": score if score > 0 else None, "Socials": socials
    }

    if check_is_existing(c_handle, item.get('id'), item.get('snippet', {}).get('customUrl'), c_title, db_existing_set):
        base_data["Lý do loại"] = "Kênh đã tồn tại trong Database CRM hoặc Giỏ hàng"
        return False, base_data
    if c_subs < min_subs_choice: base_data["Lý do loại"] = f"Dưới {min_subs_choice:,} Subs"; return False, base_data
    passes_l1, l1_reason = passes_layer1_metadata_filter(c_title, c_desc, c_country)
    if not passes_l1: base_data["Lý do loại"] = l1_reason; return False, base_data
    if c_video_count == 0 or not c_playlist or not recent_vids: base_data["Lý do loại"] = "Kênh chỉ làm Shorts / Rỗng"; return False, base_data
    if not is_within_last_90_days(latest_date): base_data["Lý do loại"] = f"Kênh ngưng hoạt động (>90 ngày, gần nhất: {latest_date})"; return False, base_data
    if not any(v.get('Seconds', 0) >= 600 for v in recent_vids): base_data["Lý do loại"] = "Không có video > 10 phút trong các video gần nhất"; return False, base_data
    return True, base_data
