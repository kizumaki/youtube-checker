import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import isodate
import pycountry
import requests
import zipfile
import os
import re
import datetime
import io
import json
import xml.etree.ElementTree as ET
from collections import Counter
from PIL import Image as PILImage
from supabase import create_client, Client
from concurrent.futures import ThreadPoolExecutor, as_completed

# Page Config
st.set_page_config(
    page_title="YT CHECKER PRO", 
    page_icon="🎙️", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Connect to Supabase
@st.cache_resource
def init_supabase():
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = init_supabase()

# --- INITIALIZE PERSISTENT STATE ---
if 'app_theme' not in st.session_state: st.session_state['app_theme'] = 'Studio Peach (Sáng)'
if 'selected_channels' not in st.session_state: st.session_state['selected_channels'] = set()
if 'api_usage' not in st.session_state: st.session_state['api_usage'] = {}
if 'chk_counter' not in st.session_state: st.session_state['chk_counter'] = 0

# Callback for Selection Sync
def toggle_select_channel(pure_handle):
    if pure_handle in st.session_state['selected_channels']:
        st.session_state['selected_channels'].remove(pure_handle)
    else:
        st.session_state['selected_channels'].add(pure_handle)

# CALLBACK: DYNAMICALLY SELECT ALL & RE-SEED CHECKBOXES
def cb_select_all(channel_list):
    for item in channel_list:
        raw_h = item.get('Handle') or item.get('handle')
        p_id = to_pure_id(raw_h)
        if p_id: st.session_state['selected_channels'].add(p_id)
    st.session_state['chk_counter'] += 1

# CALLBACK: DYNAMICALLY CLEAR ALL SELECTIONS & RE-SEED CHECKBOXES
def cb_clear_all():
    st.session_state['selected_channels'].clear()
    st.session_state['chk_counter'] += 1

def clear_selected_channels():
    cb_clear_all()

# Theme CSS Dynamic Injection & UNLOCK OVERFLOW FOR FLOATING TOOLBAR
is_dark = st.session_state['app_theme'] == 'Studio Espresso (Tối)'
bg_color = "#1E1816" if is_dark else "#F4F2F1"
card_bg = "#2A221F" if is_dark else "#FFFFFF"
text_color = "#F4F2F1" if is_dark else "#3D2F29"
border_color = "#3D2F29" if is_dark else "#E5E7EB"
sidebar_bg = "#241D1A" if is_dark else "#FFFFFF"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&display=swap');

/* Base Theme */
.stApp {{ background-color: {bg_color} !important; color: {text_color} !important; font-family: 'Montserrat', sans-serif !important; }}
section[data-testid="stSidebar"] {{ background-color: {sidebar_bg} !important; border-right: 1px solid {border_color} !important; box-shadow: 4px 0 15px rgba(0, 0, 0, 0.05) !important; }}
header[data-testid="stHeader"] {{ background-color: transparent !important; }}

/* UNLOCK INNER CONTAINERS FOR STICKY FLOATING ACTION BAR */
[data-baseweb="tab-panel"], div[data-testid="stTabPanel"], div[data-testid="stVerticalBlock"] {{
    overflow: visible !important;
}}

/* HIGH-END ARTISTIC TABS */
.stTabs [data-baseweb="tab-list"] {{ gap: 32px; background-color: transparent; padding: 0 0 4px 0; border-bottom: 2px solid #D1D5DB; }}
.stTabs [data-baseweb="tab"] {{ background-color: transparent !important; border: none !important; border-bottom: 3px solid transparent !important; border-radius: 0 !important; color: #6B7280 !important; font-weight: 700; font-size: 0.9rem; padding: 10px 4px; text-transform: uppercase; letter-spacing: 0.05em; transition: all 0.3s ease !important; cursor: pointer !important; }}
.stTabs [data-baseweb="tab"]:hover {{ color: #D95F26 !important; transform: translateY(-1px); }}
.stTabs [aria-selected="true"] {{ color: #D95F26 !important; border-bottom: 3px solid #D95F26 !important; transform: translateY(0); }}

/* Standard Card Container Styling */
div[data-testid="stVerticalBlockBorderWrapper"] {{ background-color: {card_bg} !important; border: 1px solid {border_color} !important; border-radius: 12px !important; padding: 12px !important; box-shadow: 0 2px 10px rgba(0, 0, 0, 0.03) !important; transition: transform 0.2s ease, box-shadow 0.2s ease; }}
div[data-testid="stVerticalBlockBorderWrapper"]:hover {{ box-shadow: 0 8px 24px rgba(0, 0, 0, 0.08) !important; }}

/* REAL FLOATING STICKY ACTION BAR */
div[data-testid="stVerticalBlockBorderWrapper"]:has(.sticky-action-bar) {{
    position: -webkit-sticky !important;
    position: sticky !important;
    top: 3.75rem !important;
    z-index: 9999 !important;
    background-color: {card_bg} !important;
    box-shadow: 0 10px 25px rgba(0,0,0,0.18) !important;
    border: 2px solid #D95F26 !important;
    border-top: 5px solid #D95F26 !important;
    border-radius: 12px !important;
    padding: 10px 14px !important;
    margin-bottom: 20px !important;
}}

/* ACTIVE INSPECTED CARD */
div[data-testid="stVerticalBlockBorderWrapper"]:has(div.active-card-marker) {{ border: 2px solid #D95F26 !important; box-shadow: 0 8px 24px rgba(217, 95, 38, 0.2) !important; }}
div[data-testid="stVerticalBlockBorderWrapper"]:has(div.in-cart-marker) {{ border: 2px solid #47A5D1 !important; box-shadow: 0 8px 24px rgba(71, 165, 209, 0.2) !important; }}

/* Active Banner Tag Styling */
.active-banner-tag {{ background-color: #D95F26 !important; color: #FFFFFF !important; padding: 6px 14px !important; border-radius: 8px !important; font-weight: 800 !important; margin-bottom: 12px !important; font-size: 0.85rem !important; letter-spacing: 0.05em !important; display: inline-block !important; box-shadow: 0 3px 10px rgba(217, 95, 38, 0.25) !important; }}
.active-banner-tag * {{ color: #FFFFFF !important; }}
.in-cart-banner-tag {{ background-color: #47A5D1 !important; color: #FFFFFF !important; padding: 6px 14px !important; border-radius: 8px !important; font-weight: 800 !important; margin-bottom: 12px !important; font-size: 0.85rem !important; letter-spacing: 0.05em !important; display: inline-block !important; box-shadow: 0 3px 10px rgba(71, 165, 209, 0.25) !important; }}
.in-cart-banner-tag * {{ color: #FFFFFF !important; }}

/* Inputs & Selectboxes */
.stTextInput input, .stTextArea textarea, .stSelectbox select {{ background-color: {card_bg} !important; color: {text_color} !important; border: 1px solid #D1D5DB !important; border-radius: 8px !important; font-family: 'Montserrat', sans-serif !important; }}
.stTextInput input:focus, .stTextArea textarea:focus, .stSelectbox select:focus {{ border-color: #D95F26 !important; box-shadow: 0 0 0 1px #D95F26 !important; }}

/* Default Buttons */
.stButton button {{ border-radius: 8px !important; font-weight: 700 !important; font-family: 'Montserrat', sans-serif !important; border: 1px solid #D1D5DB !important; background-color: {card_bg} !important; color: {text_color} !important; text-transform: uppercase; font-size: 0.8rem !important; letter-spacing: 0.05em; transition: all 0.25s cubic-bezier(0.16, 1, 0.3, 1) !important; }}
.stButton button:hover {{ border-color: #D95F26 !important; color: #D95F26 !important; transform: translateY(-1px) !important; }}

/* ARTISTIC PRIMARY BUTTONS */
.stButton button[kind="primary"], .stButton button[kind="primary"] *, .stButton button[kind="primary"] p, .stButton button[kind="primary"] span, .stButton button[kind="primary"] div {{ background: linear-gradient(135deg, #D95F26 0%, #E66A32 100%) !important; color: #FFFFFF !important; border: none !important; box-shadow: 0 3px 10px rgba(217, 95, 38, 0.22) !important; }}
.stButton button[kind="primary"]:hover, .stButton button[kind="primary"]:hover *, .stButton button[kind="primary"]:hover p, .stButton button[kind="primary"]:hover span, .stButton button[kind="primary"]:hover div {{ background: linear-gradient(135deg, #C24E18 0%, #D95F26 100%) !important; color: #FFFFFF !important; transform: translateY(-1px) !important; box-shadow: 0 6px 16px rgba(217, 95, 38, 0.32) !important; }}

/* Social Badges */
.social-badge {{ display: inline-block; padding: 4px 10px; border-radius: 6px; font-size: 0.75rem; font-weight: 700; margin-right: 6px; margin-bottom: 6px; text-decoration: none !important; color: #FFFFFF !important; box-shadow: 0 2px 5px rgba(0,0,0,0.15); }}
.social-email {{ background-color: #EA4335 !important; }}
.social-ig {{ background-color: #E1306C !important; }}
.social-tt {{ background-color: #000000 !important; border: 1px solid #555; }}
.social-x {{ background-color: #1DA1F2 !important; }}
.social-discord {{ background-color: #5865F2 !important; }}
.social-fb {{ background-color: #1877F2 !important; }}
.social-web {{ background-color: #10B981 !important; }}

/* Badges */
.badge-pro {{ display: inline-block; padding: 6px 12px; border-radius: 9999px; font-size: 0.7rem; font-weight: 800; text-transform: uppercase; letter-spacing: 0.08em; }}
.badge-ocean, .badge-ocean * {{ background-color: #47A5D1 !important; color: #FFFFFF !important; border: none !important; }}
.badge-score {{ padding: 4px 8px; border-radius: 6px; font-weight: 800; font-size: 0.8rem; background-color: #FFF2EB; color: #D95F26; border: 1px solid #D95F26; display: inline-block; }}
.badge-stt {{ font-weight: 800; font-size: 0.85rem; color: #6B7280; background-color: #E5E7EB; padding: 2px 8px; border-radius: 6px; margin-right: 8px; inline-block; }}
</style>
""", unsafe_allow_html=True)

# Global Default API Key
DEFAULT_API_KEY = st.secrets.get("YOUTUBE_API_KEY", "AIzaSyDDBEJscqkGGpG1xtuL4wYPuFkS4BIL854")

# --- DATABASE PERSISTENCE HELPERS ---
def load_api_keys_from_db():
    try:
        res = supabase.table("app_config").select("value").eq("key", "api_keys").execute()
        if res.data and len(res.data) > 0: return res.data[0]["value"]
    except Exception: pass
    return None

def save_api_keys_to_db(key_string):
    try:
        supabase.table("app_config").upsert({"key": "api_keys", "value": key_string}, on_conflict="key").execute()
        return True
    except Exception: return False

def load_cart_from_db():
    cart_dict = {}
    try:
        res = supabase.table("cart_items").select("*").execute()
        if res.data:
            for row in res.data:
                h = row["handle"]
                c_data = row.get("channel_data")
                if isinstance(c_data, str): c_data = json.loads(c_data)
                cart_dict[h] = c_data
    except Exception: pass
    return cart_dict

def add_to_cart_db(pure_handle, channel_data):
    try:
        data_clean = dict(channel_data)
        if "recent_videos" in data_clean: del data_clean["recent_videos"]
        supabase.table("cart_items").upsert({"handle": pure_handle, "channel_data": data_clean}, on_conflict="handle").execute()
    except Exception: pass

def remove_from_cart_db(pure_handle):
    try: supabase.table("cart_items").delete().eq("handle", pure_handle).execute()
    except Exception: pass

def clear_cart_db():
    try: supabase.table("cart_items").delete().neq("handle", "___NONE___").execute()
    except Exception: pass

def clear_entire_database():
    try:
        supabase.table("channels").delete().neq("handle", "___NONE___").execute()
        keys_to_clear = ['passed_channels', 'rejected_channels', 'batch_check_new', 'batch_check_existing', 'batch_check_rejected', 'tab5_crm_cache']
        for k in keys_to_clear:
            if k in st.session_state: del st.session_state[k]
        cb_clear_all()
        return True
    except Exception:
        return False

def load_campaigns():
    try:
        res = supabase.table("app_config").select("value").eq("key", "campaigns").execute()
        if res.data and len(res.data) > 0: return json.loads(res.data[0]["value"])
    except Exception: pass
    return {}

def save_campaigns(camps_dict):
    try: supabase.table("app_config").upsert({"key": "campaigns", "value": json.dumps(camps_dict)}, on_conflict="key").execute()
    except Exception: pass

# --- INITIALIZE PERSISTENT STATE ---
if 'global_api_keys' not in st.session_state:
    db_keys = load_api_keys_from_db()
    st.session_state['global_api_keys'] = db_keys if db_keys else DEFAULT_API_KEY

if 'cart' not in st.session_state or 'cart_loaded' not in st.session_state:
    st.session_state['cart'] = load_cart_from_db()
    st.session_state['cart_loaded'] = True

if 'pending_seed_input' in st.session_state:
    st.session_state['seed_input_tab3'] = st.session_state['pending_seed_input']
    del st.session_state['pending_seed_input']

if 'seed_input_tab3' not in st.session_state: st.session_state['seed_input_tab3'] = "@NickDiGiovanni"
if 'pending_keywords' in st.session_state:
    st.session_state['custom_kw_tab3'] = st.session_state['pending_keywords']
    del st.session_state['pending_keywords']

if 'custom_kw_tab3' not in st.session_state: st.session_state['custom_kw_tab3'] = ""
if 'video_preview_cache' not in st.session_state: st.session_state['video_preview_cache'] = {}
if 'active_inspected_handle' not in st.session_state: st.session_state['active_inspected_handle'] = None

def set_active_inspected_channel(pure_handle):
    st.session_state['active_inspected_handle'] = pure_handle

def set_api_keys(key_string):
    keys = [k.strip() for k in re.split(r'[\n,]+', key_string) if k.strip()]
    st.session_state['api_keys'] = keys if keys else [DEFAULT_API_KEY]

set_api_keys(st.session_state['global_api_keys'])

# --- CLEAN & ROBUST CUSTOM KPI RENDERER ---
def render_kpi_cards(kpi_data):
    cards_html = ""
    for title, val, color in kpi_data:
        cards_html += f'<div style="flex: 1; min-width: 200px; background-color: {card_bg}; border: 1px solid {border_color}; border-radius: 12px; padding: 18px; text-align: center; box-shadow: 0 4px 10px rgba(0,0,0,0.04);"><div style="font-size: 0.8rem; color: #6B7280; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 6px;">{title}</div><div style="font-size: 2.2rem; font-weight: 900; color: {color};">{val}</div></div>'
    full_html = f'<div style="display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap;">{cards_html}</div>'
    st.markdown(full_html, unsafe_allow_html=True)

# --- SIDEBAR BRANDING & CONFIG ---
with st.sidebar:
    col1, col2, col3 = st.columns([0.5, 8, 0.5])
    with col2:
        logo_paths = ["logo.png", "logo_2.png", "logo.jpg"]
        found_logo = None
        for path in logo_paths:
            if os.path.exists(path):
                found_logo = path
                break
        if found_logo: st.image(found_logo, use_container_width=True)
        else:
            st.markdown("""
            <div style="text-align: center; padding: 10px 0;">
                <svg width="65" height="65" viewBox="0 0 100 100" fill="none" xmlns="http://www.w3.org/2000/svg">
                    <rect x="10" y="45" width="8" height="10" rx="2" fill="#D95F26"/>
                    <rect x="22" y="30" width="8" height="40" rx="2" fill="#D95F26"/>
                    <rect x="34" y="15" width="8" height="70" rx="2" fill="#D95F26"/>
                    <rect x="46" y="5" width="8" height="90" rx="2" fill="#D95F26"/>
                    <rect x="58" y="20" width="8" height="60" rx="2" fill="#D95F26"/>
                    <rect x="70" y="35" width="8" height="30" rx="2" fill="#D95F26"/>
                    <rect x="82" y="42" width="8" height="16" rx="2" fill="#D95F26"/>
                </svg>
            </div>
            """, unsafe_allow_html=True)
            
    st.markdown("""
        <div style="text-align: center; padding-bottom: 8px;">
            <h2 style="margin: 0 0 4px 0; font-weight: 800; font-size: 1.15rem; letter-spacing: -0.02em;">YT CHECKER PRO</h2>
            <span class="badge-pro badge-ocean">Supabase Live</span>
        </div>
    """, unsafe_allow_html=True)
    
    st.selectbox("🎨 Giao diện App:", options=["Studio Peach (Sáng)", "Studio Espresso (Tối)"], key="app_theme")
    st.divider()

    # --- API HEALTH MONITOR ---
    st.markdown("<h4 style='font-weight: 700; font-size: 0.95rem;'>🛡️ Sức Khỏe API Quota</h4>", unsafe_allow_html=True)
    active_keys = st.session_state.get('api_keys', [])
    usage_data = st.session_state.get('api_usage', {})
    
    for k in active_keys:
        used = usage_data.get(k, 0)
        pct = min(100, int((used / 10000) * 100))
        color = "#10B981" if pct < 70 else ("#F59E0B" if pct < 90 else "#EF4444")
        st.markdown(f"""
            <div style='margin-bottom: 8px;'>
                <div style='font-size: 0.75rem; color: #6B7280; font-weight: 700;'>🔑 {k[:10]}...</div>
                <div style='background-color: #E5E7EB; border-radius: 4px; height: 6px; width: 100%; margin-top: 4px;'>
                    <div style='background-color: {color}; width: {pct}%; height: 100%; border-radius: 4px;'></div>
                </div>
                <div style='font-size: 0.65rem; color: #9CA3AF; text-align: right; margin-top: 2px;'>{used:,}/10,000</div>
            </div>
        """, unsafe_allow_html=True)

    keys_input = st.text_area("Cập nhật danh sách Key (1 key/dòng):", value=st.session_state['global_api_keys'], height=80, key="api_keys_text_area")
    
    if st.button("💾 Lưu Cấu Hình Key", type="primary", use_container_width=True):
        st.session_state['global_api_keys'] = keys_input
        set_api_keys(keys_input)
        save_api_keys_to_db(keys_input)
        st.toast("🎉 Đã lưu vĩnh viễn danh sách API Keys!")
        st.rerun()

    st.divider()
    if st.button("🔄 Làm Mới Màn Hình", use_container_width=True):
        cb_clear_all()
        keys_to_clear = ['passed_channels', 'rejected_channels', 'last_inspected_data', 'last_inspected_handle', 'audit_success_msg', 'batch_check_new', 'batch_check_existing', 'batch_check_rejected', 'active_inspected_handle', 'tab5_crm_cache']
        for key in keys_to_clear:
            if key in st.session_state: del st.session_state[key]
        for key in list(st.session_state.keys()):
            if key.startswith('audit_file_'): del st.session_state[key]
        st.rerun()

# --- HELPER TO DELETE CHANNEL COMPLETELY FROM SYSTEM ---
def delete_channel_from_system(pure_handle):
    if not pure_handle: return
    try: supabase.table("channels").delete().eq("handle", pure_handle).execute()
    except Exception: pass

    remove_from_cart_db(pure_handle)
    if pure_handle in st.session_state.get('cart', {}): del st.session_state['cart'][pure_handle]
    if st.session_state.get('active_inspected_handle') == pure_handle: st.session_state['active_inspected_handle'] = None
    if pure_handle in st.session_state['selected_channels']: st.session_state['selected_channels'].remove(pure_handle)

    for key_list in ['passed_channels', 'rejected_channels', 'batch_check_new', 'batch_check_existing', 'batch_check_rejected']:
        if key_list in st.session_state:
            st.session_state[key_list] = [ch for ch in st.session_state[key_list] if to_pure_id(ch.get('Handle')) != pure_handle]

    audit_key = f"audit_file_{pure_handle}"
    if audit_key in st.session_state: del st.session_state[audit_key]

# --- API QUOTA ROTATION MANAGER ---
def yt_execute(request_func, cost=1):
    keys = st.session_state.get('api_keys', [DEFAULT_API_KEY])
    if not keys: keys = [DEFAULT_API_KEY]
    idx = st.session_state.get('api_key_idx', 0)
    for attempt in range(len(keys)):
        key = keys[idx]
        try:
            yt = build("youtube", "v3", developerKey=key)
            req = request_func(yt)
            res = req.execute()
            
            usage = st.session_state.get('api_usage', {})
            usage[key] = usage.get(key, 0) + cost
            st.session_state['api_usage'] = usage
            
            return res
        except HttpError as e:
            if e.resp.status in [403, 400, 429]:
                idx = (idx + 1) % len(keys)
                st.session_state['api_key_idx'] = idx
            else: raise e
    raise Exception("❌ Toàn bộ API Keys bạn nhập đã bị chết hoặc cạn sạch Quota!")

# --- HELPER FUNCTIONS & ULTRA-DEEP CONTACT MINING ---
def to_pure_id(raw_val):
    if not raw_val or pd.isna(raw_val) or str(raw_val).strip().upper() in ["N/A", "NAN", "NONE", ""]: return None
    s = str(raw_val).strip()
    m_url = re.search(r'youtube\.com/(?:@|c/|user/|channel/)?([^\s?#/]+)', s, re.IGNORECASE)
    if m_url:
        val = m_url.group(1)
        if val.lower() in ['watch', 'shorts', 'feed', 'embed']: return None
        s = val
    s = re.sub(r'[\s]+', '', s)
    s = re.sub(r'^@+', '', s).strip().lower()

    pattern = r'_(?:backlog|january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec|\d{1,4})[_\-\s,.]?.*$'
    s = re.sub(pattern, '', s, flags=re.IGNORECASE)

    return s if s else None

# ULTRA-DEEP CONTACT & SOCIAL MEDIA EXTRACTION ENGINE
def extract_contacts_and_socials(text_corpus):
    if not text_corpus: return {}
    corpus = str(text_corpus)
    contacts = {}
    
    # 1. Emails
    emails = re.findall(r'\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b', corpus)
    valid_emails = []
    for e in emails:
        e_lower = e.lower()
        if not any(e_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp']):
            if e_lower not in ['user@domain.com', 'email@domain.com', 'yourname@email.com', 'info@youtube.com']:
                valid_emails.append(e)
    if valid_emails: contacts['Email'] = valid_emails[0]
    
    # 2. Instagram
    ig_matches = re.findall(r'(?:https?://)?(?:www\.)?(?:instagram\.com|instagr\.am)/([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if not ig_matches:
        ig_matches = re.findall(r'(?:ig|instagram)\s*[:@-]\s*@?([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if ig_matches:
        handle = ig_matches[0].rstrip('./-_,')
        if handle.lower() not in ['p', 'reel', 'reels', 'stories', 'tv', 'explore', 'direct']:
            contacts['Instagram'] = f"https://instagram.com/{handle}"
            
    # 3. TikTok
    tt_matches = re.findall(r'(?:https?://)?(?:www\.)?tiktok\.com/@?([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if not tt_matches:
        tt_matches = re.findall(r'(?:tiktok|tt)\s*[:@-]\s*@?([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if tt_matches:
        handle = tt_matches[0].rstrip('./-_,')
        contacts['TikTok'] = f"https://tiktok.com/@{handle}"

    # 4. Twitter / X
    x_matches = re.findall(r'(?:https?://)?(?:www\.)?(?:twitter\.com|x\.com)/([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if not x_matches:
        x_matches = re.findall(r'(?:twitter|x)\s*[:@-]\s*@?([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if x_matches:
        handle = x_matches[0].rstrip('./-_,')
        if handle.lower() not in ['intent', 'share', 'home', 'search', 'hashtag', 'i', 'widgets']:
            contacts['Twitter'] = f"https://x.com/{handle}"

    # 5. Discord
    dc_matches = re.findall(r'(?:https?://)?(?:www\.)?(?:discord\.(?:gg|com/invite))/([a-zA-Z0-9_-]+)', corpus, re.IGNORECASE)
    if dc_matches:
        contacts['Discord'] = f"https://discord.gg/{dc_matches[0]}"

    # 6. Facebook
    fb_matches = re.findall(r'(?:https?://)?(?:www\.)?(?:facebook\.com|fb\.com|fb\.me)/([a-zA-Z0-9_.-]+)', corpus, re.IGNORECASE)
    if fb_matches:
        handle = fb_matches[0].rstrip('./-_,')
        if handle.lower() not in ['sharer', 'groups', 'pages', 'photo', 'permalink', 'dialog']:
            contacts['Facebook'] = f"https://facebook.com/{handle}"

    # 7. Website / Linktree / Beacons
    web_matches = re.findall(r'(https?://(?:linktr\.ee|beacons\.ai|beacons\.page|bit\.ly|twitch\.tv|kick\.com|[a-zA-Z0-9-]+\.[a-zA-Z]{2,})/[^\s),"\']+)', corpus, re.IGNORECASE)
    if web_matches:
        url = web_matches[0]
        if not any(domain in url.lower() for domain in ['youtube.com', 'youtu.be', 'instagram.com', 'tiktok.com', 'twitter.com', 'x.com', 'facebook.com', 'discord.gg']):
            contacts['Website'] = url

    return contacts

def render_social_badges_html(contacts_dict):
    if not contacts_dict: return "<span style='font-size:0.75rem; color:#9CA3AF;'>Chưa có thông tin MXH</span>"
    html = "<div style='margin-top: 6px;'>"
    if 'Email' in contacts_dict: html += f"<a href='mailto:{contacts_dict['Email']}' class='social-badge social-email'>✉️ {contacts_dict['Email']}</a>"
    if 'Instagram' in contacts_dict: html += f"<a href='{contacts_dict['Instagram']}' target='_blank' class='social-badge social-ig'>📸 IG</a>"
    if 'TikTok' in contacts_dict: html += f"<a href='{contacts_dict['TikTok']}' target='_blank' class='social-badge social-tt'>🎵 TikTok</a>"
    if 'Twitter' in contacts_dict: html += f"<a href='{contacts_dict['Twitter']}' target='_blank' class='social-badge social-x'>🐦 X</a>"
    if 'Discord' in contacts_dict: html += f"<a href='{contacts_dict['Discord']}' target='_blank' class='social-badge social-discord'>💬 Discord</a>"
    if 'Facebook' in contacts_dict: html += f"<a href='{contacts_dict['Facebook']}' target='_blank' class='social-badge social-fb'>📘 FB</a>"
    if 'Website' in contacts_dict: html += f"<a href='{contacts_dict['Website']}' target='_blank' class='social-badge social-web'>🌐 Web</a>"
    html += "</div>"
    return html

# TAB 5 MULTI-THREADED CRM METADATA WORKER
def process_single_crm_channel_meta(pure_handle):
    if not pure_handle: return pure_handle, {"sub_count": -1, "sub_str": "N/A", "country": "N/A", "socials": {}}
    try:
        cid = get_channel_id_by_handle(pure_handle)
        if cid:
            playlist_id, sub_count, desc, joined, country_name, country_code, avatar = get_channel_details(cid)
            recent_vids = get_6_recent_videos(pure_handle)
            v_descs = " ".join([v.get('Description', '') for v in recent_vids]) if recent_vids else ""
            corpus = f"{desc} {v_descs}"
            socials = extract_contacts_and_socials(corpus)
            return pure_handle, {
                "sub_count": sub_count,
                "sub_str": f"{sub_count:,}" if sub_count > 0 else "N/A",
                "country": country_name if country_name else "N/A",
                "socials": socials
            }
    except Exception: pass
    return pure_handle, {"sub_count": -1, "sub_str": "N/A", "country": "N/A", "socials": {}}

@st.cache_data(ttl=86400, show_spinner=False)
def get_channel_crm_meta(pure_handle):
    _, meta = process_single_crm_channel_meta(pure_handle)
    return meta

def extract_video_id(raw_url):
    if not raw_url or pd.isna(raw_url): return None
    s = str(raw_url).strip()
    m = re.search(r'(?:v=|\/shorts\/|youtu\.be\/)([a-zA-Z0-9_-]{11})', s)
    return m.group(1) if m else None

def get_handles_from_video_ids(video_ids):
    if not video_ids: return []
    channel_ids = set()
    for i in range(0, len(video_ids), 50):
        chunk = video_ids[i:i+50]
        try:
            res = yt_execute(lambda yt: yt.videos().list(part="snippet", id=','.join(chunk)), cost=1)
            for item in res.get('items', []):
                c_id = item.get('snippet', {}).get('channelId')
                if c_id: channel_ids.add(c_id)
        except Exception: pass
        
    handles = []
    if channel_ids:
        c_ids_list = list(channel_ids)
        for i in range(0, len(c_ids_list), 50):
            chunk = c_ids_list[i:i+50]
            try:
                res = yt_execute(lambda yt: yt.channels().list(part="snippet", id=','.join(chunk)), cost=1)
                for item in res.get('items', []):
                    custom_url = item.get('snippet', {}).get('customUrl', '')
                    pure = to_pure_id(custom_url) or to_pure_id(item.get('id'))
                    if pure and pure not in handles:
                        handles.append(pure)
            except Exception: pass
    return handles

def parse_raw_inputs_to_handles(raw_inputs_list):
    handles = set()
    video_ids = set()
    for raw in raw_inputs_list:
        if not raw or pd.isna(raw): continue
        s = str(raw).strip()
        if not s: continue
        
        v_id = extract_video_id(s)
        if v_id:
            video_ids.add(v_id)
        else:
            p_h = to_pure_id(s)
            if p_h: handles.add(p_h)
            
    if video_ids:
        with st.spinner(f"🔍 Đang giải mã {len(video_ids)} Link Video sang Handle Kênh..."):
            resolved_h = get_handles_from_video_ids(list(video_ids))
            for h in resolved_h: handles.add(h)
            
    return list(handles)

def extract_raw_inputs_from_file(uploaded_file):
    raw_list = []
    fname = uploaded_file.name.lower()
    try:
        if fname.endswith('.txt'):
            content = uploaded_file.read().decode("utf-8", errors="ignore")
            raw_list = re.split(r'[\n,\t\r]+', content)
        elif fname.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
            for col in df.columns:
                for val in df[col].dropna(): raw_list.append(str(val))
        elif fname.endswith('.xlsx') or fname.endswith('.xls'):
            df = pd.read_excel(uploaded_file)
            for col in df.columns:
                for val in df[col].dropna(): raw_list.append(str(val))
    except Exception as e: st.error(f"Lỗi đọc file: {e}")
    return raw_list

def extract_handle_from_filename(filename):
    base = os.path.basename(filename)
    base_no_ext = os.path.splitext(base)[0]
    
    pattern = r'_(?:backlog|january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec|\d{1,4})[_\-\s,.]?.*$'
    cleaned = re.sub(pattern, '', base_no_ext, flags=re.IGNORECASE)
    cleaned = re.sub(r'[\s]+', '', cleaned)
    pure_id = re.sub(r'^@+', '', cleaned).strip().lower()
    return pure_id if pure_id else None

def is_long_form_video(v, min_seconds=180):
    title = v.get('Title', '').lower()
    if '#shorts' in title or '#short' in title: return False
    if v.get('Seconds', 0) <= min_seconds: return False
    return True

def is_within_last_90_days(date_str):
    if not date_str or date_str == "N/A": return False
    s = str(date_str).strip().lower()
    today = datetime.date.today()
    m_iso = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m_iso:
        try:
            dt = datetime.date(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
            return 0 <= (today - dt).days <= 90
        except Exception: return False
    return False

EXCLUDED_COUNTRIES = {'CN', 'TW', 'HK', 'TH', 'IN', 'VN'}
NON_LATIN_REGEX = re.compile(r'[\u0E00-\u0E7F]|[\u4E00-\u9FFF]|[\u0900-\u097F]|[\uAC00-\uD7AF]', re.IGNORECASE)
VIETNAMESE_UNIQUE_REGEX = re.compile(r'[ơờớởỡợưừứửữựđĐăằắẳẵặảẻỉỏủỷạẹịọụỵềếểễệồốổỗộầấẩẫậ]', re.IGNORECASE)

EXCLUDED_KEYWORDS = [
    'official mv', 'music video', 'official audio', 'album', 'song', 'records', 'lyrics', 'remix', 'vocal', 'cover', 'music', 'songs',
    'news', 'politics', 'tin tức', 'chính trị', 'thời sự', 'chiến tranh', 'đảng', 'quân sự', 'bản tin', 'điểm tin',
    'lgbt', 'lgbtq', 'gay', 'lesbian', 'transgender', 'war', 'military', 'ukraine', 'russia'
]
STOP_WORDS = {'the', 'a', 'an', 'and', 'or', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'is', 'it', 'this', 'that', 'ep', 'episode', 'part', 'video', 'shorts', 'full', 'hd', '2024', '2025', '2026', 'official', 'channel', 'vs', 'dude', 'perfect', 'nick', 'digiovanni', 'mrbeast', 'pewdiepie'}

def passes_layer1_metadata_filter(title, desc, country_code):
    if country_code in EXCLUDED_COUNTRIES: return False, f"Quốc gia bị loại ({country_code})"
    combined_text = f"{title} {desc}".lower()
    if NON_LATIN_REGEX.search(combined_text): return False, "Ngôn ngữ không phù hợp (Trung, Thái, Hindi, Hàn)"
    if VIETNAMESE_UNIQUE_REGEX.search(combined_text): return False, "Kênh Ngôn Ngữ Tiếng Việt"
    for kw in EXCLUDED_KEYWORDS:
        if kw in combined_text: return False, f"Loại nội dung cấm ({kw.upper()})"
    return True, "OK"

def clean_and_extract_keywords(text, seed_handle=""):
    seed_clean = seed_handle.replace('@', '').lower()
    words = re.findall(r'\b[a-zA-Z0-9]{3,}\b', text.lower())
    filtered = [w for w in words if w not in STOP_WORDS and w not in seed_clean]
    return filtered

# --- YOUTUBE API OPERATIONS ---
@st.cache_data(ttl=86400, show_spinner=False)
def get_channel_id_by_handle(handle):
    clean = handle.replace('@', '').split('/')[-1].strip()
    try:
        res = yt_execute(lambda yt: yt.channels().list(part="id", forHandle=clean), cost=1)
        if 'items' in res and len(res['items']) > 0: return res['items'][0]['id']
    except Exception: pass
    try:
        res = yt_execute(lambda yt: yt.search().list(part="snippet", q=clean, type="channel", maxResults=1), cost=100)
        if 'items' in res and len(res['items']) > 0: return res['items'][0]['snippet']['channelId']
    except Exception: pass
    return None

@st.cache_data(ttl=86400, show_spinner=False)
def extract_channel_master_keywords(channel_id):
    keywords_pool, channel_keywords, top_tags, categories = [], [], [], []
    try:
        ch_res = yt_execute(lambda yt: yt.channels().list(part="brandingSettings,contentDetails,snippet,topicDetails", id=channel_id), cost=1)
        if 'items' in ch_res and len(ch_res['items']) > 0:
            item = ch_res['items'][0]
            raw_kw = item.get('brandingSettings', {}).get('channel', {}).get('keywords', '')
            found_kw = re.findall(r'"([^"]+)"|\b([a-zA-Z0-9]{3,})\b', raw_kw)
            for k1, k2 in found_kw:
                kw = k1 or k2
                if kw and len(kw) > 2 and kw.lower() not in STOP_WORDS:
                    keywords_pool.append(kw.lower())
                    channel_keywords.append(kw.lower())
            topics = item.get('topicDetails', {}).get('topicCategories', [])
            for t in topics: categories.append(t.split('/')[-1].replace('_', ' '))
            uploads_playlist = item.get('contentDetails', {}).get('relatedPlaylists', {}).get('uploads', '')
            if uploads_playlist:
                v_res = yt_execute(lambda yt: yt.playlistItems().list(part="snippet", playlistId=uploads_playlist, maxResults=15), cost=1)
                v_ids = [v['snippet']['resourceId']['videoId'] for v in v_res.get('items', [])]
                if v_ids:
                    v_detail_res = yt_execute(lambda yt: yt.videos().list(part="snippet", id=','.join(v_ids)), cost=1)
                    for v_item in v_detail_res.get('items', []):
                        for tag in v_item.get('snippet', {}).get('tags', []):
                            if len(tag) > 2 and tag.lower() not in STOP_WORDS:
                                keywords_pool.append(tag.lower())
                                top_tags.append(tag.lower())
    except Exception: pass
    most_common_kws = [word for word, count in Counter(keywords_pool).most_common(15)]
    top_tag_counts = [word for word, count in Counter(top_tags).most_common(10)]
    return {"master_keywords": most_common_kws, "channel_keywords": list(set(channel_keywords))[:10], "top_tags": top_tag_counts, "categories": categories}

def get_channel_details(channel_id):
    res = yt_execute(lambda yt: yt.channels().list(part="snippet,contentDetails,statistics", id=channel_id), cost=1)
    if 'items' in res and len(res['items']) > 0:
        item = res['items'][0]
        playlist_id = item.get('contentDetails', {}).get('relatedPlaylists', {}).get('uploads', '')
        sub_count = int(item['statistics'].get('subscriberCount', 0))
        description = item['snippet'].get('description', 'No description available.')
        joined_date_raw = item['snippet'].get('publishedAt', '')
        joined_date = ""
        if joined_date_raw:
            try: joined_date = pd.to_datetime(joined_date_raw).strftime("%b %d, %Y")
            except Exception: joined_date = joined_date_raw[:10]
        country_code = item['snippet'].get('country', 'N/A')
        country_name = pycountry.countries.get(alpha_2=country_code).name if country_code != 'N/A' and pycountry.countries.get(alpha_2=country_code) else country_code
        avatar_url = item['snippet'].get('thumbnails', {}).get('high', {}).get('url', '')
        return playlist_id, sub_count, description, joined_date, country_name, country_code, avatar_url
    return None, 0, "", "", "", "", ""

def get_video_details(video_ids):
    video_data = []
    total = len(video_ids)
    if total == 0: return video_data
    for i in range(0, total, 50):
        try:
            chunk = video_ids[i:i+50]
            res = yt_execute(lambda yt: yt.videos().list(part="snippet,contentDetails,statistics", id=','.join(chunk)), cost=1)
            for item in res.get('items', []):
                duration_seconds = int(isodate.parse_duration(item['contentDetails']['duration']).total_seconds())
                h, rem = divmod(duration_seconds, 3600)
                m, s = divmod(rem, 60)
                dur_str = f"{h}:{m:02d}:{s:02d}" if h > 0 else f"{m}:{s:02d}"

                pub_date = item['snippet']['publishedAt']
                try: formatted_date = pd.to_datetime(pub_date).strftime("%d-%m-%Y")
                except Exception: formatted_date = pub_date[:10]
                
                video_data.append({
                    'Title': item['snippet']['title'], 
                    'Description': item['snippet'].get('description', ''),
                    'Link': f"https://youtube.com/watch?v={item['id']}",
                    'Length (Exact)': dur_str, 
                    'Seconds': duration_seconds,
                    'Views': int(item['statistics'].get('viewCount', 0)), 
                    'Published Date': formatted_date,
                    'Video ID': item['id']
                })
        except Exception: pass
    return video_data

# --- ULTRA-FAST 0-QUOTA RSS PREVIEW FETCH ---
@st.cache_data(ttl=43200, show_spinner=False)
def get_6_recent_videos(pure_handle):
    long_vids = []
    try:
        cid = get_channel_id_by_handle(pure_handle)
        if cid:
            rss_url = f"https://www.youtube.com/feeds/videos.xml?channel_id={cid}"
            resp = requests.get(rss_url, timeout=4)
            if resp.status_code == 200:
                root = ET.fromstring(resp.content)
                ns = {'atom': 'http://www.w3.org/2005/Atom', 'yt': 'http://www.youtube.com/xml/schemas/2015'}
                rss_v_ids = []
                for entry in root.findall('atom:entry', ns):
                    v_id_el = entry.find('yt:videoId', ns)
                    if v_id_el is not None and v_id_el.text: rss_v_ids.append(v_id_el.text)
                    if len(rss_v_ids) >= 12: break
                if rss_v_ids:
                    v_details = get_video_details(rss_v_ids)
                    for v in v_details:
                        if is_long_form_video(v, min_seconds=180): long_vids.append(v)
                        if len(long_vids) >= 6: break

            if len(long_vids) < 6:
                playlist_id, _, _, _, _, _, _ = get_channel_details(cid)
                if playlist_id:
                    v_res = yt_execute(lambda yt: yt.playlistItems().list(part="snippet", playlistId=playlist_id, maxResults=30), cost=1)
                    v_ids = [v_item['snippet']['resourceId']['videoId'] for v_item in v_res.get('items', [])]
                    if v_ids:
                        v_details = get_video_details(v_ids)
                        for v in v_details:
                            if is_long_form_video(v, min_seconds=180) and v not in long_vids: long_vids.append(v)
                            if len(long_vids) >= 6: break
    except Exception: pass
    return long_vids[:6]

# --- AI OUTREACH EMAIL DRAFT MODAL ---
@st.dialog("🤖 AI Soạn Mail Hợp Tác", width="large")
def show_ai_email_dialog(channel_data):
    st.markdown(f"<h3 style='color: #D95F26; font-weight: 800;'>Thư ngỏ gửi <span style='color: #47A5D1;'>{channel_data.get('Tên Kênh', 'Creator')}</span></h3>", unsafe_allow_html=True)
    st.caption("✨ AI đã tự động phân tích chỉ số và thiết kế thư cá nhân hóa. Bạn có thể chỉnh sửa trước khi Copy.")
    
    er_text = channel_data.get('ER', 'rất ấn tượng')
    if er_text != 'N/A' and er_text != 'rất ấn tượng': er_text = f"lên tới {er_text}"
        
    template = f"""Subject: Collaboration Opportunity with {channel_data.get('Tên Kênh', 'your channel')} 🚀

Hi {channel_data.get('Tên Kênh', 'there')},

I’ve been closely following your content on YouTube, especially your latest videos, and I am absolutely blown away by the quality! It’s no surprise your channel is growing so fast, with an incredible engagement rate {er_text}. 

I am reaching out from Backstreet Voice Studio. We specialize in high-end voiceover, localization, and audio optimization, helping top-tier creators like you expand their reach into new global markets effortlessly. 

Given your strong audience retention and niche focus, translating and dubbing your content could easily double your current viewership without requiring any extra production effort on your end.

Would you be open to a quick 10-minute chat this week to see how we could partner up to maximize your channel's revenue? 

Keep up the amazing work!

Best regards,
[Your Name/Title]
Backstreet Voice Studio"""
    
    st.text_area("Bản Thảo Email (Sẵn sàng Copy):", value=template, height=350)
    if st.button("❌ Đóng Cửa Sổ", type="primary", use_container_width=True): st.rerun()

# --- CONFIRM CLEAR ENTIRE DATABASE DIALOG ---
@st.dialog("⚠️ CẢNH BÁO: XÓA SẠCH DATABASE", width="small")
def confirm_clear_db_dialog():
    st.error("🚨 Hành động này sẽ XÓA VĨNH VIỄN toàn bộ danh sách kênh trong Supabase và KHÔNG THỂ HỒI PHỤC!")
    st.write("Vui lòng gõ **`XOA DATABASE`** vào ô bên dưới để xác nhận:")
    confirm_txt = st.text_input("Xác nhận:", key="input_confirm_db_wipe")
    if st.button("💣 XÁC NHẬN XÓA SẠCH DATABASE", type="primary", use_container_width=True):
        if confirm_txt.strip().upper() == "XOA DATABASE":
            if clear_entire_database():
                st.success("🎉 Đã xóa sạch vĩnh viễn toàn bộ Database!")
                st.rerun()
            else: st.error("❌ Đã xảy ra lỗi khi kết nối Supabase!")
        else:
            st.warning("⚠️ Mã xác nhận không đúng! Vui lòng gõ 'XOA DATABASE'.")

# --- STREAMLIT MODAL DIALOGS ---
@st.dialog("🎬 6 Video Dài (Long-form) Mới Nhất", width="large")
def show_video_dialog(pure_handle, pre_fetched_videos=None):
    st.markdown(f"<h3 style='font-weight: 800;'>📺 Kênh đang xem: <span style='color: #D95F26;'>@{pure_handle}</span></h3>", unsafe_allow_html=True)
    st.markdown(f"🔗 **[Mở thẳng Tab Videos trên YouTube](https://youtube.com/@{pure_handle}/videos)**")
    
    vids = []
    if pre_fetched_videos: vids = [v for v in pre_fetched_videos if is_long_form_video(v, min_seconds=180)]
    if len(vids) < 6: vids = get_6_recent_videos(pure_handle)
    vids = vids[:6]
    
    if vids:
        st.divider()
        for row_idx in range(0, len(vids), 2):
            col1, col2 = st.columns(2)
            v1 = vids[row_idx]
            with col1:
                vid_id1 = v1.get('Video ID') or (v1.get('Link', '').split('v=')[-1] if 'v=' in v1.get('Link', '') else '')
                if vid_id1: st.image(f"https://img.youtube.com/vi/{vid_id1}/hqdefault.jpg", use_container_width=True)
                st.markdown(f"**[{v1['Title'][:45]}...]({v1['Link']})**")
                st.caption(f"👀 {v1.get('Views', 0):,} views | ⏳ {v1.get('Length (Exact)', 'N/A')} | 📅 {v1.get('Published Date', '')}")
            
            if row_idx + 1 < len(vids):
                v2 = vids[row_idx + 1]
                with col2:
                    vid_id2 = v2.get('Video ID') or (v2.get('Link', '').split('v=')[-1] if 'v=' in v2.get('Link', '') else '')
                    if vid_id2: st.image(f"https://img.youtube.com/vi/{vid_id2}/hqdefault.jpg", use_container_width=True)
                    st.markdown(f"**[{v2['Title'][:45]}...]({v2['Link']})**")
                    st.caption(f"👀 {v2.get('Views', 0):,} views | ⏳ {v2.get('Length (Exact)', 'N/A')} | 📅 {v2.get('Published Date', '')}")
            st.divider()
    else: st.caption("Không tìm thấy video dài (Kênh này chỉ đăng Shorts hoặc chưa có video dài trên 3 phút).")
    if st.button("❌ Đóng Cửa Sổ Preview", type="primary", use_container_width=True): st.rerun()

def get_selected_channel_data():
    selected = st.session_state['selected_channels']
    data = []
    lists = st.session_state.get('batch_check_new', []) + st.session_state.get('batch_check_existing', []) + st.session_state.get('passed_channels', []) + st.session_state.get('rejected_channels', []) + st.session_state.get('batch_check_rejected', [])
    seen = set()
    for item in lists:
        p = to_pure_id(item.get('Handle'))
        if p in selected and p not in seen:
            data.append(item)
            seen.add(p)
    return data

@st.dialog("⚖️ Bảng So Sánh Kênh Trực Diện", width="large")
def compare_channels_dialog(channel_data_list):
    if not channel_data_list:
        st.warning("Không có dữ liệu kênh để so sánh.")
        return
    st.markdown("<h3 style='text-align: center; color: #D95F26; font-weight: 800; margin-bottom: 20px;'>📊 SO SÁNH CHỈ SỐ KÊNH</h3>", unsafe_allow_html=True)
    
    enriched_list = []
    with st.spinner("Đang kết nối API cào dữ liệu chi tiết để so sánh..."):
        for ch in channel_data_list:
            c_dict = dict(ch)
            pure_h = to_pure_id(c_dict.get('Handle'))
            if pure_h:
                if c_dict.get('Tổng Số Video') is None or c_dict.get('Tổng Số Video') == 'N/A' or c_dict.get('Video Gần Nhất') is None or c_dict.get('Video Gần Nhất') == 'N/A':
                    cid = get_channel_id_by_handle(pure_h)
                    if cid:
                        playlist_id, sub_count, channel_desc, channel_joined, country_name, country_code, avatar_url = get_channel_details(cid)
                        recent_vids = get_6_recent_videos(pure_h)
                        latest_date = recent_vids[0]['Published Date'] if recent_vids else 'N/A'
                        try:
                            c_res = yt_execute(lambda yt: yt.channels().list(part="statistics", id=cid), cost=1)
                            video_count = int(c_res['items'][0]['statistics'].get('videoCount', 0)) if (c_res.get('items') and len(c_res['items']) > 0) else 0
                        except Exception: video_count = 0
                        
                        if not c_dict.get('Subscribers') or c_dict.get('Subscribers') == 'N/A':
                            c_dict['Subscribers'] = f"{sub_count:,}"
                        c_dict['Tổng Số Video'] = f"{video_count:,}"
                        if not c_dict.get('Quốc gia') or c_dict.get('Quốc gia') == 'N/A':
                            c_dict['Quốc gia'] = country_name if country_name else 'N/A'
                        c_dict['Video Gần Nhất'] = latest_date
                        
                        if recent_vids and sub_count > 0:
                            avg_views = sum(v.get('Views', 0) for v in recent_vids) / len(recent_vids)
                            er_rate = (avg_views / sub_count) * 100
                            c_dict['ER'] = f"{er_rate:.2f}%"
                            c_dict['Score'] = min(100, int((er_rate / 10.0) * 100))
            enriched_list.append(c_dict)

    cols = st.columns(len(enriched_list))
    for idx, ch in enumerate(enriched_list):
        with cols[idx]:
            ch_handle = ch.get('Handle', 'N/A')
            pure_h = to_pure_id(ch_handle)
            ch_link = f"https://youtube.com/@{pure_h}" if pure_h else "javascript:void(0);"
            score_badge = f"<div style='margin-top: 10px;'><span class='badge-score'>🔥 Điểm: {ch.get('Score', 'N/A')}/100</span></div>" if ch.get('Score') else ""

            card_html = f"""
            <div style='background-color: {card_bg}; padding: 18px 12px; border-radius: 12px; border: 1px solid {border_color}; text-align: center; margin-bottom: 10px;'>
                <h4 style='color: #47A5D1; font-weight: 800; margin-bottom: 5px;'><a href='{ch_link}' style='text-decoration: none; color: inherit;'>{ch_handle}</a></h4>
                <p style='font-size: 0.85rem; color: #6B7280; font-weight: 600; margin-bottom: 12px;'>{ch.get('Tên Kênh', 'N/A')}</p>
                <hr style='border: none; border-top: 1px solid {border_color}; margin: 10px 0;'>
                <p style='font-size: 0.75rem; color: #6B7280; margin-bottom: 2px; font-weight: 700;'>👥 SUBSCRIBERS</p>
                <p style='font-size: 1.4rem; font-weight: 800; color: #D95F26; margin-top: 0; margin-bottom: 12px;'>{ch.get('Subscribers', 'N/A')}</p>
                <p style='font-size: 0.75rem; color: #6B7280; margin-bottom: 2px; font-weight: 700;'>🎬 TỔNG VIDEO</p>
                <p style='font-size: 1.2rem; font-weight: 700; color: {text_color}; margin-top: 0; margin-bottom: 12px;'>{ch.get('Tổng Số Video', 'N/A')}</p>
                <p style='font-size: 0.75rem; color: #6B7280; margin-bottom: 2px; font-weight: 700;'>🌍 QUỐC GIA</p>
                <p style='font-size: 1.0rem; font-weight: 600; color: {text_color}; margin-top: 0; margin-bottom: 12px;'>{ch.get('Quốc gia', 'N/A')}</p>
                <p style='font-size: 0.75rem; color: #6B7280; margin-bottom: 2px; font-weight: 700;'>📅 GẦN NHẤT</p>
                <p style='font-size: 1.0rem; font-weight: 600; color: #47A5D1; margin-top: 0; margin-bottom: 0;'>{ch.get('Video Gần Nhất', 'N/A')}</p>
                {score_badge}
            </div>
            """
            st.markdown(card_html, unsafe_allow_html=True)

    st.write("")
    if st.button("❌ Đóng Cửa Sổ So Sánh", type="primary", use_container_width=True): st.rerun()

def run_single_channel_audit(pure_handle):
    cid = get_channel_id_by_handle(pure_handle)
    if not cid: return None, None
    playlist_id, sub_count, channel_desc, channel_joined, channel_country, c_code, avatar_url = get_channel_details(cid)
    v_ids = []
    next_token = None
    while True:
        res = yt_execute(lambda yt: yt.playlistItems().list(part="snippet", playlistId=playlist_id, maxResults=50, pageToken=next_token), cost=1)
        for v_item in res.get('items', []): v_ids.append(v_item['snippet']['resourceId']['videoId'])
        next_token = res.get('nextPageToken')
        if not next_token: break
    v_data = get_video_details(v_ids)
    excel_bytes = generate_v414_excel_report(pure_handle, sub_count, channel_desc, channel_joined, channel_country, avatar_url, v_data)
    out_fname = f"{pure_handle}_{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx"
    return excel_bytes, out_fname

def generate_v414_excel_report(clean_handle, sub_count, channel_desc, channel_joined, channel_country, avatar_url, video_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = clean_handle[:31]
    date_str = datetime.datetime.now().strftime("%d-%m-%Y")
    total_videos = len(video_data)
    total_views = sum(v['Views'] for v in video_data)
    total_minutes = round(sum(v['Seconds'] for v in video_data) / 60)
    ws.merge_cells('A1:E1')
    ws['A1'] = f"{clean_handle.upper()} YOUTUBE CHANNEL SUMMARY REPORT - up to {date_str}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="D95F26", end_color="D95F26", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws['A2'] = f"Total Videos: {total_videos:,}"
    ws['A3'] = f"Total Duration: {total_minutes:,} minutes"
    ws['A4'] = f"Total Views: {total_views:,}"
    ws['A5'] = f"Total Subscribers: {sub_count:,}"
    ws['A6'] = f"Country Location: {channel_country}"
    ws['A7'] = f"Channel Joined Date: {channel_joined}"
    ws['A9'] = "Channel Description Text:"
    ws['A9'].font = Font(bold=True, italic=True)
    ws['A10'] = channel_desc
    ws['A10'].alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(2, 8): ws[f'A{row}'].font = Font(bold=True)
    if avatar_url:
        try:
            res = requests.get(avatar_url, timeout=5)
            if res.status_code == 200:
                img = PILImage.open(io.BytesIO(res.content))
                img = img.resize((140, 140))
                temp_buf = io.BytesIO()
                img.save(temp_buf, format="PNG")
                temp_buf.seek(0)
                ws.add_image(ExcelImage(temp_buf), 'C10')
        except Exception: pass
    headers = ["Video Title", "Link", "Length", "Views", "Published Date"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=12, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 24
    for idx, v in enumerate(video_data):
        r = idx + 13
        cA = ws.cell(row=r, column=1, value=v['Title']); cA.font = Font(name="Calibri", size=11)
        cB = ws.cell(row=r, column=2, value=v['Link'])
        if v.get('Link'): cB.hyperlink = v['Link']; cB.font = Font(name="Calibri", size=11, color="47A5D1", underline="single")
        ws.cell(row=r, column=3, value=v['Length (Exact)']).alignment = Alignment(horizontal="center", vertical="center")
        cD = ws.cell(row=r, column=4, value=v['Views']); cD.number_format = '#,##0'
        ws.cell(row=r, column=5, value=v['Published Date']).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions['A'].width = 55; ws.column_dimensions['B'].width = 45; ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 15; ws.column_dimensions['E'].width = 15
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# --- REUSABLE COMPONENT: RENDER SHARED CART ---
def render_shared_cart_ui(key_suffix=""):
    st.divider()
    cart_items = st.session_state['cart']
    
    col_t1, col_t2 = st.columns([7, 3])
    with col_t1: st.markdown(f"<h3 style='font-weight: 800;'>🛒 Giỏ Hàng Dùng Chung ({len(cart_items)} Kênh)</h3>", unsafe_allow_html=True)
    with col_t2:
        with st.expander("📁 Quản Lý Chiến Dịch"):
            camps = load_campaigns()
            new_camp_name = st.text_input("Tên chiến dịch mới:", key=f"new_camp_name_{key_suffix}")
            if st.button("💾 Lưu Giỏ Hàng Thành Chiến Dịch", use_container_width=True, key=f"save_camp_btn_{key_suffix}"):
                if new_camp_name:
                    camps[new_camp_name] = cart_items
                    save_campaigns(camps)
                    st.success(f"Đã lưu chiến dịch '{new_camp_name}'!")
                else: st.warning("Vui lòng nhập tên chiến dịch.")
            
            sel_camp = st.selectbox("Tải lại chiến dịch cũ:", options=["-- Chọn --"] + list(camps.keys()), key=f"sel_camp_{key_suffix}")
            if st.button("📂 Tải Dữ Liệu", use_container_width=True, key=f"load_camp_btn_{key_suffix}"):
                if sel_camp != "-- Chọn --":
                    st.session_state['cart'] = camps[sel_camp]
                    clear_cart_db()
                    for k, v in st.session_state['cart'].items(): add_to_cart_db(k, v)
                    st.success(f"Đã tải thành công chiến dịch {sel_camp}!")
                    st.rerun()

    if cart_items:
        df_cart = pd.DataFrame(list(cart_items.values()))
        if 'Handle' in df_cart.columns:
            df_cart['Tab Videos'] = df_cart['Handle'].apply(lambda h: f"https://youtube.com/@{to_pure_id(h)}/videos" if to_pure_id(h) else "")
            df_cart['Link Kênh'] = df_cart['Handle'].apply(lambda h: f"https://youtube.com/@{to_pure_id(h)}" if to_pure_id(h) else "")
            
        if 'recent_videos' in df_cart.columns: df_cart = df_cart.drop(columns=['recent_videos'])

        df_cart.index = range(1, len(df_cart) + 1)

        st.dataframe(df_cart, use_container_width=True, column_config={
            "Link Kênh": st.column_config.LinkColumn("Trang Chủ", display_text="🏠 Kênh"),
            "Tab Videos": st.column_config.LinkColumn("Tab Videos", display_text="🎬 Videos"),
            "Tag": st.column_config.TextColumn("🏷️ Nhãn Trạng Thái")
        })
        
        with st.expander("🚀 Đẩy Dữ Liệu & Xuất File (Google Sheets / Excel)"):
            c1, c2, c3, c4 = st.columns(4)
            c1.download_button("📄 Tải TXT", data="\n".join([i["Handle"] for i in cart_items.values()]), file_name="gio_hang_dung_chung.txt", use_container_width=True, key=f"dl_txt_cart_{key_suffix}")
            buf_xl = io.BytesIO(); df_cart.to_excel(buf_xl, index=False)
            c2.download_button("📊 Tải Excel", data=buf_xl.getvalue(), file_name="gio_hang_dung_chung.xlsx", use_container_width=True, key=f"dl_xl_cart_{key_suffix}")
            if c3.button("⚡ Nạp Toàn Bộ Vào DB", type="primary", use_container_width=True, key=f"push_db_cart_{key_suffix}"):
                data_db = [{"handle": to_pure_id(i["Handle"]), "youtuber_name": i.get("Tên Kênh", ""), "source": f"Cart Import [{i.get('Tag', '')}]"} for i in cart_items.values()]
                supabase.table("channels").upsert(data_db, on_conflict="handle").execute()
                st.success(f"🎉 Đã nạp {len(data_db)} kênh vào Database!")
            if c4.button("🧹 Xóa Sạch Giỏ Hàng", use_container_width=True, key=f"clear_cart_{key_suffix}"): 
                clear_cart_db()
                st.session_state['cart'] = {}
                st.success("🎉 Đã xóa sạch Giỏ Hàng!")
                st.rerun()

            st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
            wh_col1, wh_col2 = st.columns([3, 1])
            with wh_col1: webhook_url = st.text_input("🔗 Dán Google Apps Script Webhook URL (Để đồng bộ lên Sheets):", key=f"webhook_url_{key_suffix}")
            with wh_col2: 
                st.write("")
                if st.button("🚀 Push to Google Sheets", type="primary", use_container_width=True, key=f"push_gsheets_btn_{key_suffix}"):
                    if webhook_url:
                        try:
                            requests.post(webhook_url, json={"data": df_cart.to_dict(orient="records")})
                            st.success("🎉 Đã đẩy dữ liệu thành công!")
                        except Exception as e: st.error(f"Lỗi: {e}")
                    else: st.warning("Vui lòng dán Webhook URL!")
    else:
        st.info("Giỏ hàng đang trống. Bấm '🛒 Thêm' ở Tab 1 hoặc Tab 3 để nhặt kênh vào giỏ!")

# --- APP HEADER ---
st.markdown("""
    <div style="padding: 5px 0 15px 0;">
        <h1 style="font-weight: 900; margin-bottom: 5px; font-size: 2.4rem; letter-spacing: -0.03em;">YT CHECKER <span style="color: #D95F26;">PRO</span></h1>
        <p style="font-size: 1.05rem; font-weight: 500; opacity: 0.8;">Hệ thống phân tích, tìm kiếm kênh đồng ngách Đa Luồng Siêu Tốc & Quản lý Chiến Dịch.</p>
    </div>
""", unsafe_allow_html=True)

# Tabs
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "🔍 Tra cứu Handle Hàng Loạt", 
    "⚡ Cào Live & Tạo Báo Cáo Audit", 
    "🎯 Săn Kênh Tương Tự (Multi-Threaded)",
    "📤 Upload Cập nhật Data", 
    "📊 Xem Database",
    "✨ Soi Từ Khóa Kênh (SEO Inspector)"
])

# --- HELPER FOR IN-MEMORY SORTING & FILTERING ---
def sort_and_filter_channels(channel_list, search_query, sort_by):
    filtered = list(channel_list)
    if search_query:
        q = search_query.strip().lower()
        filtered = [c for c in filtered if q in c.get('Handle', '').lower() or q in c.get('Tên Kênh', '').lower()]
    if sort_by == "Subscribers (Cao -> Thấp)":
        def parse_subs(val):
            try: return int(str(val.get('Subscribers', '0')).replace(',', ''))
            except Exception: return 0
        filtered.sort(key=parse_subs, reverse=True)
    elif sort_by == "Tên Kênh (A -> Z)": filtered.sort(key=lambda x: x.get('Tên Kênh', '').lower())
    elif sort_by == "Mới Đăng Video": filtered.sort(key=lambda x: x.get('Video Gần Nhất', ''), reverse=True)
    return filtered

# --- DATABASE-FIRST WORKER FOR TAB 1 ---
def process_tab1_single_handle(p_id, db_matches):
    if p_id in db_matches:
        db_item = db_matches[p_id]
        return "EXISTING", {
            "Handle": f"@{p_id}",
            "Tên Kênh": db_item.get("youtuber_name", p_id.upper()),
            "Trạng thái": "❌ Đã có trong DB"
        }
    
    cid = get_channel_id_by_handle(p_id)
    if not cid:
        return "REJECTED", {
            "Handle": f"@{p_id}",
            "Tên Kênh": p_id.upper(),
            "Trạng thái": "❌ Không tìm thấy kênh",
            "Lý do loại": "Không tồn tại trên YT"
        }
    
    try:
        res = yt_execute(lambda yt: yt.channels().list(part="snippet,statistics", id=cid), cost=1)
        if res.get('items'):
            item = res['items'][0]
            ch_title = item['snippet'].get('title', p_id.upper())
            ch_desc = item['snippet'].get('description', '')
            country_code = item['snippet'].get('country', 'N/A')
            country_name = pycountry.countries.get(alpha_2=country_code).name if country_code != 'N/A' and pycountry.countries.get(alpha_2=country_code) else country_code
            sub_count = int(item['statistics'].get('subscriberCount', 0))
            
            # STEP 1: DEEP SOCIAL CONTACT MINING WITH VIDEO DESCRIPTIONS
            recent_vids = get_6_recent_videos(p_id)
            v_descs = " ".join([v.get('Description', '') for v in recent_vids])
            combined_text_corpus = f"{ch_title} {ch_desc} {v_descs}"
            social_contacts = extract_contacts_and_socials(combined_text_corpus)
            
            # FILTER 1: MUST HAVE >= 1,000,000 SUBS
            if sub_count < 1000000:
                return "REJECTED", {
                    "Handle": f"@{p_id}",
                    "Tên Kênh": ch_title,
                    "Subscribers": f"{sub_count:,}",
                    "Trạng thái": f"❌ Dưới 1 triệu Subs ({sub_count:,})",
                    "Lý do loại": f"Dưới 1M Subs ({sub_count:,})",
                    "Socials": social_contacts
                }
            
            passes_l1, l1_reason = passes_layer1_metadata_filter(ch_title, ch_desc, country_code)
            if not passes_l1:
                return "REJECTED", {
                    "Handle": f"@{p_id}",
                    "Tên Kênh": ch_title,
                    "Subscribers": f"{sub_count:,}",
                    "Trạng thái": f"❌ {l1_reason}",
                    "Lý do loại": l1_reason,
                    "Socials": social_contacts
                }
            
            return "NEW", {
                "Handle": f"@{p_id}",
                "Tên Kênh": ch_title,
                "Subscribers": f"{sub_count:,}",
                "Quốc gia": country_name,
                "Link Kênh": f"https://www.youtube.com/@{p_id}",
                "Trạng thái": "✅ Kênh Mới Đạt Chuẩn",
                "Socials": social_contacts
            }
    except Exception: pass
    
    return "REJECTED", {
        "Handle": f"@{p_id}",
        "Tên Kênh": p_id.upper(),
        "Trạng thái": "❌ Lỗi đọc dữ liệu API",
        "Lý do loại": "Lỗi API",
        "Socials": {}
    }

# --- TAB 1: BATCH SEARCH WITH STICKY FLOATING TOOLBAR & PAGE COUNTER ---
with tab1:
    st.markdown("<h3 style='font-weight: 700; margin-top: 15px;'>🔍 Kiểm tra Trùng Lặp Danh Sách Handle / Link Video Hàng Loạt</h3>", unsafe_allow_html=True)
    st.caption("💡 *Tự động quét quy mô kênh (Yêu cầu >= 1,000,000 Subs), lọc bỏ các kênh Âm nhạc/News/LGBT và Đào sâu MXH/Email.*")
    
    col_s1, col_s2 = st.columns([2, 1])
    with col_s1: text_input_area = st.text_area("Dán danh sách Handle/Link kênh/Link Video vào đây (mỗi dòng 1 link):", height=180)
    with col_s2: file_input_check = st.file_uploader("Hoặc Upload file danh sách (.txt, .csv, .xlsx):")
        
    if st.button("🔎 Bắt Đầu Kiểm Tra Hàng Loạt", type="primary"):
        all_raw_inputs = []
        if text_input_area:
            all_raw_inputs.extend(re.split(r'[\n,\t\r]+', str(text_input_area)))
        if file_input_check:
            all_raw_inputs.extend(extract_raw_inputs_from_file(file_input_check))
                
        target_list = parse_raw_inputs_to_handles(all_raw_inputs)
        if not target_list: st.warning("⚠️ Vui lòng dán danh sách Handle, Link Video hoặc chọn file để kiểm tra!")
        else:
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            response = supabase.table("channels").select("handle, youtuber_name").in_("handle", target_list).execute()
            db_matches = {item["handle"].lower(): item for item in response.data} if response.data else {}
            
            new_handles, existing_handles, rejected_handles = [], [], []
            total_items = len(target_list)
            completed_count = 0
            
            with ThreadPoolExecutor(max_workers=8) as executor:
                futures = [executor.submit(process_tab1_single_handle, p_id, db_matches) for p_id in target_list]
                for future in as_completed(futures):
                    status, res_data = future.result()
                    if status == "NEW": new_handles.append(res_data)
                    elif status == "EXISTING": existing_handles.append(res_data)
                    else: rejected_handles.append(res_data)
                    
                    completed_count += 1
                    progress_bar.progress(completed_count / total_items)
                    status_text.markdown(f"⏳ **Đang phân tích siêu tốc:** `{completed_count}/{total_items}` Handle...")

            progress_bar.empty()
            status_text.empty()

            st.session_state['batch_check_new'] = new_handles
            st.session_state['batch_check_existing'] = existing_handles
            st.session_state['batch_check_rejected'] = rejected_handles

    if 'batch_check_new' in st.session_state or 'batch_check_existing' in st.session_state or 'batch_check_rejected' in st.session_state:
        new_handles = st.session_state.get('batch_check_new', [])
        existing_handles = st.session_state.get('batch_check_existing', [])
        rejected_handles = st.session_state.get('batch_check_rejected', [])

        st.divider()
        render_kpi_cards([
            ("TỔNG SỐ KIỂM TRA", f"{len(new_handles) + len(existing_handles) + len(rejected_handles)}", "#47A5D1"),
            ("✅ ĐẠT CHUẨN (>=1M SUBS)", f"{len(new_handles)}", "#10B981"),
            ("❌ ĐÃ TỒN TẠI TRONG DB", f"{len(existing_handles)}", "#F59E0B"),
            ("🚫 BỊ LOẠI (<1M / CẤM)", f"{len(rejected_handles)}", "#EF4444")
        ])
        
        res_tab1, res_tab2, res_tab3 = st.tabs([
            f"✅ Kênh Mới Đạt Chuẩn ({len(new_handles)})", 
            f"❌ Kênh Đã Tồn Tại ({len(existing_handles)})",
            f"🚫 Kênh Bị Loại ({len(rejected_handles)})"
        ])
        
        with res_tab1:
            if new_handles:
                cart_keys = set(st.session_state['cart'].keys())
                selected_set = st.session_state['selected_channels']
                
                selected_not_in_cart = [p for p in selected_set if p not in cart_keys]
                cnt_for_cart = len(selected_not_in_cart)
                cnt_total_sel = len(selected_set)

                # SAFE STICKY FLOATING ACTION BAR
                with st.container(border=True):
                    st.markdown('<div class="sticky-action-bar"></div>', unsafe_allow_html=True)
                    tb1, tb2, tb3, tb4, tb5, tb6 = st.columns([2.0, 2.0, 1.8, 1.8, 1.2, 1.2])
                    with tb1:
                        if st.button("🛒 Thêm TẤT CẢ Vào Giỏ", type="primary", key="btn_add_all_t1", use_container_width=True):
                            for item in new_handles:
                                p_id = to_pure_id(item["Handle"])
                                item_data = {"Handle": item["Handle"], "Tên Kênh": item.get("Tên Kênh", p_id.upper()), "Link Kênh": f"https://www.youtube.com/@{p_id}", "Trạng Thái DB": "✅ KÊNH MỚI", "Tag": "📌 Chưa phân loại", "Socials": item.get("Socials", {})}
                                st.session_state['cart'][p_id] = item_data
                                add_to_cart_db(p_id, item_data)
                            st.success(f"🎉 Đã thêm {len(new_handles)} kênh mới vào Giỏ hàng chung!")
                            st.rerun()
                    with tb2:
                        if st.button(f"🛒 Thêm ({cnt_for_cart}) Đã Chọn", key="btn_add_sel_t1", use_container_width=True):
                            if cnt_for_cart > 0:
                                for item in new_handles:
                                    p_id = to_pure_id(item["Handle"])
                                    if p_id in selected_not_in_cart:
                                        item_data = {"Handle": item["Handle"], "Tên Kênh": item.get("Tên Kênh", p_id.upper()), "Link Kênh": f"https://www.youtube.com/@{p_id}", "Trạng Thái DB": "✅ KÊNH MỚI", "Tag": "📌 Chưa phân loại", "Socials": item.get("Socials", {})}
                                        st.session_state['cart'][p_id] = item_data
                                        add_to_cart_db(p_id, item_data)
                                st.success(f"🎉 Đã thêm {cnt_for_cart} kênh mới vào giỏ!")
                                st.rerun()
                            else: st.warning("Không có kênh mới nào chưa được thêm vào giỏ trong các kênh bạn chọn!")
                    with tb3:
                        if st.button(f"⚖️ So Sánh ({cnt_total_sel}) Kênh", key="btn_cmp_sel_t1", use_container_width=True):
                            if 1 < cnt_total_sel <= 5: compare_channels_dialog(get_selected_channel_data())
                            else: st.warning("Vui lòng chọn từ 2 đến 5 kênh để so sánh!")
                    with tb4:
                        if st.button(f"🗑️ Xóa ({cnt_total_sel}) Đã Chọn", key="btn_del_sel_t1", use_container_width=True):
                            if cnt_total_sel > 0:
                                for p_id in list(selected_set): delete_channel_from_system(p_id)
                                st.session_state['selected_channels'].clear()
                                st.success(f"🗑️ Đã xóa {cnt_total_sel} kênh!")
                                st.rerun()
                            else: st.warning("Vui lòng tick chọn ít nhất 1 kênh!")
                    with tb5:
                        st.button("✅ Chọn Tất Cả", key="btn_sel_all_t1_new", on_click=cb_select_all, args=(new_handles,), use_container_width=True)
                    with tb6:
                        st.button("❌ Bỏ Chọn", key="btn_clear_sel_t1_new", on_click=cb_clear_all, use_container_width=True)

                items_per_page = 20
                total_pages = max(1, (len(new_handles) + items_per_page - 1) // items_per_page)
                
                col_p1, col_p2 = st.columns([2, 8])
                with col_p1:
                    page_new = st.number_input("Trang (Kênh Mới):", min_value=1, max_value=total_pages, value=1, step=1, key="page_new_t1")
                with col_p2:
                    st.write("")
                    st.markdown(f"📄 **Trang {page_new} / {total_pages}** *(Hiển thị {min(20, len(new_handles))} / {len(new_handles)} kênh)*")

                start_idx = (page_new - 1) * items_per_page
                paged_new = new_handles[start_idx:start_idx + items_per_page]

                for idx, item in enumerate(paged_new):
                    p_id = to_pure_id(item["Handle"])
                    is_active = (p_id == st.session_state.get('active_inspected_handle'))
                    is_in_cart = p_id in st.session_state['cart']
                    stt_num = start_idx + idx + 1
                    
                    with st.container(border=True):
                        if is_active:
                            st.markdown('<div class="active-card-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="active-banner-tag">🔍 ĐANG XEM 6 VIDEO MỚI CỦA KÊNH NÀY</div>', unsafe_allow_html=True)
                        elif is_in_cart:
                            st.markdown('<div class="in-cart-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="in-cart-banner-tag">🛒 ĐÃ CÓ TRONG GIỎ HÀNG</div>', unsafe_allow_html=True)

                        c0, c1, c2, c3 = st.columns([0.4, 3.1, 3.5, 3.0])
                        with c0:
                            st.checkbox("", key=f"chk_t1_{p_id}_{st.session_state['chk_counter']}", value=(p_id in st.session_state['selected_channels']), on_change=toggle_select_channel, args=(p_id,))
                        with c1:
                            st.markdown(f"<h3 style='margin:0; font-weight:800; font-size:1.3rem;'><span class='badge-stt'>#{stt_num}</span><a href='{item['Link Kênh']}' style='color:#D95F26; text-decoration:none;'>{item['Handle']}</a></h3>", unsafe_allow_html=True)
                            st.write(f"**{item.get('Tên Kênh', p_id.upper())}**")
                            
                            c1_1, c1_2 = st.columns(2)
                            if c1_1.button("👁️ Xem Video", key=f"btn_prev_t1_{p_id}", on_click=set_active_inspected_channel, args=(p_id,)):
                                show_video_dialog(p_id)
                            if c1_2.button("📩 Soạn Mail", key=f"btn_mail_t1_{p_id}"):
                                show_ai_email_dialog(item)

                        with c2:
                            st.write(f"👥 **Subs:** `{item.get('Subscribers', 'N/A')}` | 🌍 **Q.Gia:** `{item.get('Quốc gia', 'N/A')}`")
                            st.markdown(f"**Trạng thái:** <span style='color:#10B981; font-weight:700;'>{item['Trạng thái']}</span>", unsafe_allow_html=True)
                            st.markdown(render_social_badges_html(item.get("Socials", {})), unsafe_allow_html=True)
                        with c3:
                            st.write("**Thao tác:**")
                            bc1, bc2 = st.columns(2)
                            if is_in_cart:
                                current_tag = st.session_state['cart'][p_id].get("Tag", "📌 Chưa phân loại")
                                new_tag = st.selectbox("Gắn Nhãn:", ["📌 Chưa phân loại", "🔥 Ưu tiên làm", "📩 Đã liên hệ", "⏳ Đang chờ duyệt", "✅ Đã chốt", "❌ Bỏ qua"], index=["📌 Chưa phân loại", "🔥 Ưu tiên làm", "📩 Đã liên hệ", "⏳ Đang chờ duyệt", "✅ Đã chốt", "❌ Bỏ qua"].index(current_tag), key=f"tag_t1_{p_id}")
                                if new_tag != current_tag:
                                    st.session_state['cart'][p_id]["Tag"] = new_tag
                                    add_to_cart_db(p_id, st.session_state['cart'][p_id])
                                if bc1.button("❌ Bỏ Giỏ", key=f"rm_t1_{p_id}", use_container_width=True):
                                    remove_from_cart_db(p_id)
                                    del st.session_state['cart'][p_id]
                                    st.rerun()
                            else:
                                if bc1.button("🛒 Thêm Giỏ", key=f"add_t1_{p_id}", use_container_width=True):
                                    item_data = {"Handle": item["Handle"], "Tên Kênh": item.get("Tên Kênh", p_id.upper()), "Link Kênh": f"https://www.youtube.com/@{p_id}", "Trạng Thái DB": "✅ KÊNH MỚI", "Tag": "📌 Chưa phân loại", "Socials": item.get("Socials", {})}
                                    st.session_state['cart'][p_id] = item_data
                                    add_to_cart_db(p_id, item_data)
                                    st.rerun()

                            if bc2.button("🗑️ Xóa Kênh", key=f"del_t1_{p_id}", use_container_width=True):
                                delete_channel_from_system(p_id)
                                st.toast(f"🗑️ Đã xóa kênh @{p_id}!")
                                st.rerun()
            else:
                st.info("Không có kênh mới nào đạt chuẩn!")

        with res_tab2:
            if existing_handles:
                with st.container(border=True):
                    st.markdown('<div class="sticky-action-bar"></div>', unsafe_allow_html=True)
                    cnt_total_sel_ex = len(st.session_state['selected_channels'])
                    te1, te2, te3 = st.columns([3, 1.5, 1.5])
                    with te1:
                        if st.button(f"🗑️ Xóa ({cnt_total_sel_ex}) Kênh Đã Chọn", key="btn_del_sel_t1_ext", use_container_width=True):
                            if cnt_total_sel_ex > 0:
                                for p_id in list(st.session_state['selected_channels']): delete_channel_from_system(p_id)
                                cb_clear_all()
                                st.success(f"🗑️ Đã xóa {cnt_total_sel_ex} kênh!")
                                st.rerun()
                            else: st.warning("Vui lòng tick chọn ít nhất 1 kênh!")
                    with te2:
                        st.button("✅ Chọn Tất Cả", key="btn_sel_all_t1_ext", on_click=cb_select_all, args=(existing_handles,), use_container_width=True)
                    with te3:
                        st.button("❌ Bỏ Chọn Tất Cả", key="btn_clear_sel_t1_ext", on_click=cb_clear_all, use_container_width=True)

                items_per_page_ex = 20
                total_pages_ex = max(1, (len(existing_handles) + items_per_page_ex - 1) // items_per_page_ex)
                
                col_pe1, col_pe2 = st.columns([2, 8])
                with col_pe1:
                    page_ex = st.number_input("Trang (Kênh Tồn Tại):", min_value=1, max_value=total_pages_ex, value=1, step=1, key="page_ex_t1")
                with col_pe2:
                    st.write("")
                    st.markdown(f"📄 **Trang {page_ex} / {total_pages_ex}** *(Hiển thị {min(20, len(existing_handles))} / {len(existing_handles)} kênh)*")

                start_idx_ex = (page_ex - 1) * items_per_page_ex
                paged_ex = existing_handles[start_idx_ex:start_idx_ex + items_per_page_ex]

                for idx, item in enumerate(paged_ex):
                    p_id = to_pure_id(item["Handle"])
                    is_active = (p_id == st.session_state.get('active_inspected_handle'))
                    stt_num_ex = start_idx_ex + idx + 1
                    
                    with st.container(border=True):
                        if is_active:
                            st.markdown('<div class="active-card-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="active-banner-tag">🔍 ĐANG XEM 6 VIDEO MỚI CỦA KÊNH NÀY</div>', unsafe_allow_html=True)

                        c0, c1, c2, c3 = st.columns([0.4, 3.6, 4.0, 2.0])
                        with c0:
                            st.checkbox("", key=f"chk_t1_ext_{p_id}_{st.session_state['chk_counter']}", value=(p_id in st.session_state['selected_channels']), on_change=toggle_select_channel, args=(p_id,))
                        with c1:
                            st.markdown(f"<h3 style='margin:0; font-weight:800; font-size:1.3rem;'><span class='badge-stt'>#{stt_num_ex}</span><a href='https://youtube.com/@{p_id}' style='text-decoration:none;'>{item['Handle']}</a></h3>", unsafe_allow_html=True)
                            st.write(f"**{item.get('Tên Kênh', 'N/A')}**")
                            if st.button("👁️ Xem 6 Video Mới", key=f"btn_prev_t1_ext_{p_id}", on_click=set_active_inspected_channel, args=(p_id,)):
                                show_video_dialog(p_id)
                        with c2:
                            st.markdown(f"**Trạng thái:** <span style='color:#F59E0B; font-weight:700;'>{item['Trạng thái']}</span>", unsafe_allow_html=True)
                        with c3:
                            if st.button("🗑️ Xóa DB", key=f"del_t1_ext_{p_id}", use_container_width=True):
                                delete_channel_from_system(p_id)
                                st.toast(f"🗑️ Đã xóa kênh @{p_id} khỏi DB!")
                                st.rerun()

        with res_tab3:
            if rejected_handles:
                with st.container(border=True):
                    st.markdown('<div class="sticky-action-bar"></div>', unsafe_allow_html=True)
                    cnt_total_sel_rej = len(st.session_state['selected_channels'])
                    tr1, tr2, tr3 = st.columns([3, 1.5, 1.5])
                    with tr1:
                        if st.button(f"🗑️ Xóa ({cnt_total_sel_rej}) Kênh Đã Chọn", key="btn_del_sel_t1_rej", use_container_width=True):
                            if cnt_total_sel_rej > 0:
                                for p_id in list(st.session_state['selected_channels']): delete_channel_from_system(p_id)
                                cb_clear_all()
                                st.success(f"🗑️ Đã xóa {cnt_total_sel_rej} kênh!")
                                st.rerun()
                            else: st.warning("Vui lòng tick chọn ít nhất 1 kênh!")
                    with tr2:
                        st.button("✅ Chọn Tất Cả", key="btn_sel_all_t1_rej", on_click=cb_select_all, args=(rejected_handles,), use_container_width=True)
                    with tr3:
                        st.button("❌ Bỏ Chọn", key="btn_clear_sel_t1_rej", on_click=cb_clear_all, use_container_width=True)

                items_per_page_rej = 20
                total_pages_rej = max(1, (len(rejected_handles) + items_per_page_rej - 1) // items_per_page_rej)
                
                col_pr1, col_pr2 = st.columns([2, 8])
                with col_pr1:
                    page_rej = st.number_input("Trang (Kênh Bị Loại):", min_value=1, max_value=total_pages_rej, value=1, step=1, key="page_rej_t1")
                with col_pr2:
                    st.write("")
                    st.markdown(f"📄 **Trang {page_rej} / {total_pages_rej}** *(Hiển thị {min(20, len(rejected_handles))} / {len(rejected_handles)} kênh)*")

                start_idx_rej = (page_rej - 1) * items_per_page_rej
                paged_rejected = rejected_handles[start_idx_rej:start_idx_rej + items_per_page_rej]

                for idx, item in enumerate(paged_rejected):
                    p_id = to_pure_id(item["Handle"])
                    is_active = (p_id == st.session_state.get('active_inspected_handle'))
                    stt_num_rej = start_idx_rej + idx + 1
                    
                    with st.container(border=True):
                        if is_active:
                            st.markdown('<div class="active-card-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="active-banner-tag">🔍 ĐANG XEM 6 VIDEO MỚI CỦA KÊNH NÀY</div>', unsafe_allow_html=True)

                        c0, c1, c2, c3 = st.columns([0.4, 3.6, 4.0, 2.0])
                        with c0:
                            st.checkbox("", key=f"chk_t1_rej_{p_id}_{st.session_state['chk_counter']}", value=(p_id in st.session_state['selected_channels']), on_change=toggle_select_channel, args=(p_id,))
                        with c1:
                            st.markdown(f"<h3 style='margin:0; font-weight:800; font-size:1.3rem;'><span class='badge-stt'>#{stt_num_rej}</span><a href='https://youtube.com/@{p_id}' style='text-decoration:none;'>{item['Handle']}</a></h3>", unsafe_allow_html=True)
                            st.write(f"**{item.get('Tên Kênh', 'N/A')}**")
                            if st.button("👁️ Xem 6 Video Mới", key=f"btn_prev_t1_rej_{p_id}", on_click=set_active_inspected_channel, args=(p_id,)):
                                show_video_dialog(p_id)
                        with c2:
                            st.write(f"👥 **Subs:** `{item.get('Subscribers', 'N/A')}`")
                            st.markdown(f"❌ **Lý do loại:** <span style='color:#EF4444; font-weight:700;'>{item.get('Lý do loại', item['Trạng thái'])}</span>", unsafe_allow_html=True)
                        with c3:
                            if st.button("🗑️ Xóa Kênh", key=f"del_t1_rej_{p_id}", use_container_width=True):
                                delete_channel_from_system(p_id)
                                st.toast(f"🗑️ Đã xóa kênh @{p_id}!")
                                st.rerun()

    render_shared_cart_ui(key_suffix="tab1")

# --- TAB 2: LIVE API SCRAPER ---
with tab2:
    st.markdown("<h3 style='font-weight: 700; margin-top: 15px;'>⚡ Cào dữ liệu Live & Xuất Báo Cáo Audit chuẩn V4.14</h3>", unsafe_allow_html=True)
    channel_url_input = st.text_input("Dán Link kênh hoặc Handle vào đây:", value="@4wd247")

    if channel_url_input and st.button("🚀 Xử lý Kênh & Tạo Báo Cáo V4.14", type="primary"):
        pure_h = to_pure_id(channel_url_input)
        if pure_h:
            try:
                b_data, f_name = run_single_channel_audit(pure_h)
                if b_data:
                    supabase.table("channels").upsert([{"handle": pure_h, "youtuber_name": pure_h.upper(), "source": "YouTube API V4.14"}], on_conflict="handle").execute()
                    st.success(f"🎉 Đã dựng xong báo cáo Audit!")
                    st.download_button("📥 Tải về File Audit V4.14", data=b_data, file_name=f_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            except Exception as e: st.error(f"Lỗi: {e}")

# --- TAB 3: MULTI-THREADED SMART RELATED FINDER ---
def process_single_candidate(item, min_subs_choice, min_duration_choice, db_existing_set):
    c_handle = to_pure_id(item['snippet'].get('customUrl', '')) or item['id'].lower()
    c_title = item['snippet']['title']
    c_desc = item['snippet'].get('description', '')
    c_country = item['snippet'].get('country', 'N/A')
    c_subs = int(item['statistics'].get('subscriberCount', 0))
    c_video_count = int(item['statistics'].get('videoCount', 0))
    c_url = f"https://www.youtube.com/@{c_handle}"
    db_status = "❌ Đã có trong DB" if c_handle in db_existing_set else "✅ KÊNH MỚI"
    
    c_playlist = item.get('contentDetails', {}).get('relatedPlaylists', {}).get('uploads', '')
    latest_date, has_qualifying_video = "N/A", False
    recent_vids = []
    avg_views, er_rate, score = 0, 0, 0
    
    if c_playlist and c_video_count > 0:
        try:
            v_res = yt_execute(lambda yt: yt.playlistItems().list(part="snippet", playlistId=c_playlist, maxResults=50), cost=1)
            v_ids = [v_item['snippet']['resourceId']['videoId'] for v_item in v_res.get('items', [])]
            if v_ids:
                v_details = get_video_details(v_ids)
                long_vids = [v for v in v_details if is_long_form_video(v, min_seconds=180)]
                if long_vids:
                    latest_date = long_vids[0]['Published Date']
                    has_qualifying_video = any(v['Seconds'] >= min_duration_choice for v in long_vids)
                    recent_vids = long_vids[:6]
                    
                    if recent_vids and c_subs > 0:
                        avg_views = sum(v.get('Views', 0) for v in recent_vids) / len(recent_vids)
                        er_rate = (avg_views / c_subs) * 100
                        score = min(100, int((er_rate / 10.0) * 100))
        except Exception: pass

    # STEP 1 CONTACT EXTRACTION WITH VIDEO DESCRIPTIONS
    v_descs = " ".join([v.get('Description', '') for v in recent_vids])
    combined_corpus = f"{c_title} {c_desc} {v_descs}"
    social_contacts = extract_contacts_and_socials(combined_corpus)

    base_data = {
        "Handle": f"@{c_handle}", "Link Kênh": c_url, "Tên Kênh": c_title, 
        "Subscribers": f"{c_subs:,}", "Quốc gia": c_country, 
        "Video Gần Nhất": latest_date, "Tổng Số Video": f"{c_video_count:,}", 
        "Trạng Thái DB": db_status, "recent_videos": recent_vids,
        "ER": f"{er_rate:.2f}%" if er_rate > 0 else "N/A",
        "Score": score if score > 0 else None,
        "Socials": social_contacts
    }

    if c_subs < min_subs_choice: 
        base_data["Lý do loại"] = f"Dưới {min_subs_choice:,} Subs"
        return False, base_data
    passes_l1, l1_reason = passes_layer1_metadata_filter(c_title, c_desc, c_country)
    if not passes_l1: 
        base_data["Lý do loại"] = l1_reason
        return False, base_data
    if c_video_count == 0 or not c_playlist: 
        base_data["Lý do loại"] = "Kênh trống"
        return False, base_data
    if not is_within_last_90_days(latest_date): 
        base_data["Lý do loại"] = f"Bỏ trống (Mới nhất: {latest_date})"
        return False, base_data
    if not has_qualifying_video: 
        base_data["Lý do loại"] = "Shorts-only"
        return False, base_data
        
    return True, base_data

with tab3:
    st.markdown("<h3 style='font-weight: 700; margin-top: 15px;'>🎯 Săn Kênh Tương Tự & Giỏ Hàng (Multi-threaded Speed)</h3>", unsafe_allow_html=True)
    
    if 'audit_success_msg' in st.session_state:
        st.success(st.session_state['audit_success_msg'])
        del st.session_state['audit_success_msg']

    trigger_auto_start_search = False
    if st.session_state.get('trigger_deep_search_now', False):
        st.session_state['trigger_deep_search_now'] = False
        trigger_auto_start_search = True

    col_f1, col_f2 = st.columns([2, 1])
    with col_f1:
        seed_channel_input = st.text_input("Nhập Handle Kênh Mồi (ví dụ: @NickDiGiovanni):", key="seed_input_tab3")
        
        if st.button("✨ Tự Động Phân Tích từ Kênh Mồi"):
            pure_s_auto = to_pure_id(seed_channel_input)
            if pure_s_auto:
                try:
                    cid_auto = get_channel_id_by_handle(pure_s_auto)
                    if cid_auto:
                        ext = extract_channel_master_keywords(cid_auto)
                        st.session_state['pending_keywords'] = ", ".join(ext['master_keywords'][:6])
                        st.rerun()
                except Exception as e: st.error(f"Lỗi: {e}")
                    
        custom_keywords_input = st.text_input("Từ khóa chủ đề (Tự động liên kết):", key="custom_kw_tab3")
        
    with col_f2:
        min_subs_choice = st.selectbox("Mốc Subscribers Tối Thiểu:", options=[100000, 250000, 500000, 1000000], index=3, format_func=lambda x: f"{x:,} Subs")
        min_duration_choice = st.selectbox("Lọc Loại Bỏ Kênh Shorts:", options=[60, 180, 300, 600], index=0, format_func=lambda x: f"Loại Shorts < {x//60} phút" if x < 600 else "Có Video > 10 phút")

    start_btn = st.button("🚀 Bắt Đầu Săn Kênh Đồng Ngách", type="primary")

    if (start_btn or trigger_auto_start_search) and seed_channel_input:
        pure_seed = to_pure_id(seed_channel_input)
        try:
            st.info(f"🔍 Đang kết nối API và phân tích `{pure_seed}`...")
            seed_id = get_channel_id_by_handle(pure_seed)
            if not seed_id: st.error("Không tìm thấy kênh mồi này trên YouTube!")
            else:
                playlist_id, _, seed_desc, _, _, _, _ = get_channel_details(seed_id)
                
                if custom_keywords_input: top_kw_list = clean_and_extract_keywords(custom_keywords_input, seed_handle=pure_seed)
                else:
                    ext_info = extract_channel_master_keywords(seed_id)
                    top_kw_list = ext_info['master_keywords'][:4] if ext_info['master_keywords'] else [pure_seed.replace('_', ' ')]
                    
                st.write(f"🏷️ **Từ khóa quét:** `{', '.join(top_kw_list)}`")
                
                candidate_channel_ids = set()
                q_chan = " ".join(top_kw_list[:2])
                c_search_res = yt_execute(lambda yt: yt.search().list(part="snippet", q=q_chan, type="channel", maxResults=50), cost=100)
                for c_item in c_search_res.get('items', []):
                    if c_item['snippet']['channelId'] != seed_id: candidate_channel_ids.add(c_item['snippet']['channelId'])
                    
                search_queries = [" ".join(top_kw_list[:2]), " ".join(top_kw_list[2:4])] if len(top_kw_list) >= 4 else [" ".join(top_kw_list)]
                for q in search_queries:
                    if not q.strip(): continue
                    v_search_res = yt_execute(lambda yt: yt.search().list(part="snippet", q=q, type="video", maxResults=50), cost=100)
                    for v_item in v_search_res.get('items', []):
                        if v_item['snippet']['channelId'] != seed_id: candidate_channel_ids.add(v_item['snippet']['channelId'])
                        
                candidate_ids_list = list(candidate_channel_ids)
                
                if not candidate_ids_list: st.warning("Không quét được ứng viên nào!")
                else:
                    passed_channels, rejected_channels = [], []
                    channel_items, candidate_handles = [], []
                    
                    for i in range(0, len(candidate_ids_list), 50):
                        chan_res = yt_execute(lambda yt: yt.channels().list(part="snippet,contentDetails,statistics", id=','.join(candidate_ids_list[i:i+50])), cost=1)
                        for item in chan_res.get('items', []):
                            c_h = to_pure_id(item['snippet'].get('customUrl', '')) or item['id'].lower()
                            candidate_handles.append(c_h)
                            channel_items.append(item)

                    db_res = supabase.table("channels").select("handle").in_("handle", candidate_handles).execute()
                    db_existing_set = {r["handle"].lower() for r in db_res.data} if db_res.data else set()
                    
                    progress_bar_t3 = st.progress(0)
                    status_text_t3 = st.empty()
                    tot_cand = len(channel_items)
                    comp_cand = 0

                    with ThreadPoolExecutor(max_workers=8) as executor:
                        futures = [executor.submit(process_single_candidate, item, min_subs_choice, min_duration_choice, db_existing_set) for item in channel_items]
                        for future in as_completed(futures):
                            is_pass, res_data = future.result()
                            if is_pass: passed_channels.append(res_data)
                            else: rejected_channels.append(res_data)
                            
                            comp_cand += 1
                            progress_bar_t3.progress(comp_cand / tot_cand)
                            status_text_t3.markdown(f"📊 **Đang phân tích siêu tốc & Đào sâu MXH:** `{comp_cand}/{tot_cand}` ứng viên...")

                    progress_bar_t3.empty()
                    status_text_t3.empty()

                    st.session_state['passed_channels'] = passed_channels
                    st.session_state['rejected_channels'] = rejected_channels

        except Exception as e: st.error(f"Lỗi: {e}")

    if 'passed_channels' in st.session_state or 'rejected_channels' in st.session_state:
        passed_list = st.session_state.get('passed_channels', [])
        rejected_list = st.session_state.get('rejected_channels', [])
        
        st.divider()
        render_kpi_cards([
            ("TỔNG ỨNG VIÊN", f"{len(passed_list) + len(rejected_list)}", "#47A5D1"),
            (f"✅ ĐẠT CHUẨN (>{min_subs_choice:,} SUBS)", f"{len(passed_list)}", "#10B981"),
            ("❌ BỊ LOẠI", f"{len(rejected_list)}", "#EF4444")
        ])

        tab_pass, tab_rej = st.tabs([f"✅ Kênh Đạt Chuẩn ({len(passed_list)})", f"❌ Kênh Bị Loại ({len(rejected_list)})"])
        
        # --- TAB PASSED ---
        with tab_pass:
            if passed_list:
                sf_col1, sf_col2, sf_col3 = st.columns([3, 2, 2])
                with sf_col1: filter_q = st.text_input("⚡ Lọc nhanh tên/handle:", key="filter_pass_q", placeholder="Gõ tên kênh để lọc...")
                with sf_col2: sort_by = st.selectbox("Sắp xếp danh sách:", options=["Subscribers (Cao -> Thấp)", "Tên Kênh (A -> Z)", "Mới Đăng Video"], key="sort_pass_by")
                with sf_col3:
                    st.write(""); st.write("")
                    if st.button("🛒 Thêm TẤT CẢ Kênh Mới Vào Giỏ", type="primary", use_container_width=True):
                        for row in passed_list:
                            if "✅" in row["Trạng Thái DB"]: 
                                p_id = to_pure_id(row["Handle"])
                                item_data = dict(row); item_data["Tag"] = "📌 Chưa phân loại"
                                st.session_state['cart'][p_id] = item_data
                                add_to_cart_db(p_id, item_data)
                        st.success("🎉 Đã thêm tất cả vào giỏ hàng!")
                        st.rerun()

                display_passed = sort_and_filter_channels(passed_list, filter_q, sort_by)

                cart_keys = set(st.session_state['cart'].keys())
                selected_set = st.session_state['selected_channels']
                
                selected_not_in_cart = [p for p in selected_set if p not in cart_keys]
                cnt_for_cart = len(selected_not_in_cart)
                cnt_total_sel = len(selected_set)

                # STICKY FLOATING ACTION BAR
                with st.container(border=True):
                    st.markdown('<div class="sticky-action-bar"></div>', unsafe_allow_html=True)
                    ba1, ba2, ba3, ba4, ba5 = st.columns([2.5, 2.5, 2.0, 1.5, 1.5])
                    with ba1:
                        if st.button(f"🛒 Thêm ({cnt_for_cart}) Kênh Mới Vào Giỏ", key="btn_add_sel_pass", use_container_width=True):
                            if cnt_for_cart > 0:
                                for row in display_passed:
                                    p_id = to_pure_id(row['Handle'])
                                    if p_id in selected_not_in_cart:
                                        item_data = dict(row); item_data["Tag"] = "📌 Chưa phân loại"
                                        st.session_state['cart'][p_id] = item_data
                                        add_to_cart_db(p_id, item_data)
                                st.success(f"🎉 Đã thêm {cnt_for_cart} kênh mới đã chọn!")
                                st.rerun()
                            else: st.warning("Không có kênh mới nào chưa được thêm vào giỏ trong các kênh bạn chọn!")
                    with ba2:
                        if st.button(f"⚖️ So Sánh ({cnt_total_sel}) Kênh", key="btn_cmp_sel_pass", use_container_width=True):
                            if 1 < cnt_total_sel <= 5: compare_channels_dialog(get_selected_channel_data())
                            else: st.warning("Vui lòng chọn từ 2 đến 5 kênh để so sánh!")
                    with ba3:
                        if st.button(f"🗑️ Xóa ({cnt_total_sel}) Đã Chọn", key="btn_del_sel_pass", use_container_width=True):
                            if cnt_total_sel > 0:
                                for p_id in list(selected_set): delete_channel_from_system(p_id)
                                cb_clear_all()
                                st.success(f"🗑️ Đã xóa {cnt_total_sel} kênh!")
                                st.rerun()
                            else: st.warning("Vui lòng tick chọn ít nhất 1 kênh!")
                    with ba4:
                        st.button("✅ Chọn Tất Cả", key="btn_sel_all_pass_t3", on_click=cb_select_all, args=(display_passed,), use_container_width=True)
                    with ba5:
                        st.button("❌ Bỏ Chọn", key="btn_clear_sel_pass", on_click=cb_clear_all, use_container_width=True)

                items_per_page = 20
                total_pages = max(1, (len(display_passed) + items_per_page - 1) // items_per_page)
                
                col_pp1, col_pp2 = st.columns([2, 8])
                with col_pp1:
                    page_pass = st.number_input("Trang (Kênh Đạt Chuẩn):", min_value=1, max_value=total_pages, value=1, step=1, key="page_pass_t3")
                with col_pp2:
                    st.write("")
                    st.markdown(f"📄 **Trang {page_pass} / {total_pages}** *(Hiển thị {min(20, len(display_passed))} / {len(display_passed)} kênh)*")

                start_idx = (page_pass - 1) * items_per_page
                paged_passed = display_passed[start_idx:start_idx + items_per_page]

                for idx, row in enumerate(paged_passed):
                    p_id = to_pure_id(row['Handle'])
                    is_active = (p_id == st.session_state.get('active_inspected_handle'))
                    is_in_cart = p_id in st.session_state['cart']
                    stt_num_pass = start_idx + idx + 1

                    with st.container(border=True):
                        if is_active:
                            st.markdown('<div class="active-card-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="active-banner-tag">🔍 ĐANG XEM 6 VIDEO MỚI CỦA KÊNH NÀY</div>', unsafe_allow_html=True)
                        elif is_in_cart:
                            st.markdown('<div class="in-cart-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="in-cart-banner-tag">🛒 ĐÃ CÓ TRONG GIỎ HÀNG</div>', unsafe_allow_html=True)

                        c0, c1, c2, c3, c4 = st.columns([0.4, 2.2, 3.0, 1.8, 3.0])
                        with c0:
                            st.checkbox("", key=f"chk_p_{p_id}_{st.session_state['chk_counter']}", value=(p_id in st.session_state['selected_channels']), on_change=toggle_select_channel, args=(p_id,))
                        with c1:
                            st.markdown(f"<h3 style='margin:0; font-weight:800; font-size:1.3rem;'><span class='badge-stt'>#{stt_num_pass}</span><a href='{row['Link Kênh']}' style='color:#D95F26; text-decoration:none;'>{row['Handle']}</a></h3>", unsafe_allow_html=True)
                            st.write(f"**{row['Tên Kênh']}**")
                            
                            c1_1, c1_2 = st.columns(2)
                            if c1_1.button("👁️ Xem Video", key=f"btn_prev_pass_{p_id}", on_click=set_active_inspected_channel, args=(p_id,)):
                                show_video_dialog(p_id, pre_fetched_videos=row.get('recent_videos'))
                            if c1_2.button("📩 Soạn Mail", key=f"btn_mail_pass_{p_id}"):
                                show_ai_email_dialog(row)

                        with c2:
                            st.write(f"👥 **Subs:** `{row['Subscribers']}` | 🌍 **Q.Gia:** `{row['Quốc gia']}`")
                            st.write(f"🎬 **Tổng Video:** `{row['Tổng Số Video']}` | 📅 **Mới nhất:** `{row['Video Gần Nhất']}`")
                            if row.get('Score'):
                                st.markdown(f"<span class='badge-score'>🔥 Điểm tiềm năng: {row['Score']}/100</span>", unsafe_allow_html=True)
                            st.markdown(render_social_badges_html(row.get("Socials", {})), unsafe_allow_html=True)
                        with c3:
                            st.write(f"**Database:**\n<span style='color:#47A5D1; font-weight:700;'>{row['Trạng Thái DB']}</span>", unsafe_allow_html=True)
                        with c4:
                            bc1, bc2 = st.columns(2)
                            if is_in_cart:
                                current_tag = st.session_state['cart'][p_id].get("Tag", "📌 Chưa phân loại")
                                new_tag = st.selectbox("Gắn Nhãn:", ["📌 Chưa phân loại", "🔥 Ưu tiên làm", "📩 Đã liên hệ", "⏳ Đang chờ duyệt", "✅ Đã chốt", "❌ Bỏ qua"], index=["📌 Chưa phân loại", "🔥 Ưu tiên làm", "📩 Đã liên hệ", "⏳ Đang chờ duyệt", "✅ Đã chốt", "❌ Bỏ qua"].index(current_tag), key=f"tag_p_{p_id}")
                                if new_tag != current_tag:
                                    st.session_state['cart'][p_id]["Tag"] = new_tag
                                    add_to_cart_db(p_id, st.session_state['cart'][p_id])
                                if bc1.button("❌ Bỏ Giỏ", key=f"rm_p_{p_id}", use_container_width=True):
                                    remove_from_cart_db(p_id)
                                    del st.session_state['cart'][p_id]
                                    st.rerun()
                            else:
                                if bc1.button("🛒 Thêm Giỏ", key=f"add_p_{p_id}", use_container_width=True):
                                    item_data = dict(row); item_data["Tag"] = "📌 Chưa phân loại"
                                    st.session_state['cart'][p_id] = item_data
                                    add_to_cart_db(p_id, item_data)
                                    st.rerun()
                                    
                            audit_key = f"audit_file_{p_id}"
                            if audit_key in st.session_state:
                                bc2.download_button("📥 Tải Audit", data=st.session_state[audit_key]["bytes"], file_name=st.session_state[audit_key]["filename"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_p_{p_id}", use_container_width=True)
                            else:
                                if bc2.button("📄 Tạo Audit", key=f"btn_p_{p_id}", use_container_width=True):
                                    with st.spinner("Đang dựng Audit..."):
                                        b_data, f_name = run_single_channel_audit(p_id)
                                        if b_data:
                                            supabase.table("channels").upsert([{"handle": p_id, "youtuber_name": row['Tên Kênh'], "source": "Smart Finder Audit"}], on_conflict="handle").execute()
                                            st.session_state['audit_success_msg'] = f"🎉 Đã lưu **@{p_id}** vào Database!"
                                            st.session_state[audit_key] = {"bytes": b_data, "filename": f_name}
                                            st.rerun()

                            bc3, bc4 = st.columns(2)
                            if bc3.button("🎯 Đào Sâu", key=f"deep_p_{p_id}", type="secondary", use_container_width=True):
                                cid_deep = get_channel_id_by_handle(p_id)
                                if cid_deep:
                                    ext_deep = extract_channel_master_keywords(cid_deep)
                                    st.session_state['pending_keywords'] = ", ".join(ext_deep['master_keywords'][:6])
                                    st.session_state['pending_seed_input'] = f"@{p_id}"
                                    st.session_state['trigger_deep_search_now'] = True
                                    st.rerun()

                            if bc4.button("🗑️ Xóa Kênh", key=f"del_p_{p_id}", use_container_width=True):
                                delete_channel_from_system(p_id)
                                st.toast(f"🗑️ Đã xóa kênh @{p_id}!")
                                st.rerun()
            else:
                st.info("Không có kênh nào đạt chuẩn.")
                
        # --- TAB REJECTED ---
        with tab_rej:
            if rejected_list:
                rf_col1, rf_col2 = st.columns([3, 3])
                with rf_col1: filter_rej_q = st.text_input("⚡ Lọc nhanh kênh bị loại:", key="filter_rej_q", placeholder="Gõ tên/handle...")
                with rf_col2: sort_rej_by = st.selectbox("Sắp xếp:", options=["Subscribers (Cao -> Thấp)", "Tên Kênh (A -> Z)"], key="sort_rej_by")

                display_rejected = sort_and_filter_channels(rejected_list, filter_rej_q, sort_rej_by)
                
                with st.container(border=True):
                    st.markdown('<div class="sticky-action-bar"></div>', unsafe_allow_html=True)
                    cnt_total_sel_t3_rej = len(st.session_state['selected_channels'])
                    ra1, ra2, ra3 = st.columns([3, 1.5, 1.5])
                    with ra1:
                        if st.button(f"🗑️ Xóa ({cnt_total_sel_t3_rej}) Đã Chọn", key="btn_del_sel_t3_rej", use_container_width=True):
                            if cnt_total_sel_t3_rej > 0:
                                for p_id in list(st.session_state['selected_channels']): delete_channel_from_system(p_id)
                                cb_clear_all()
                                st.success(f"🗑️ Đã xóa {cnt_total_sel_t3_rej} kênh!")
                                st.rerun()
                            else: st.warning("Vui lòng tick chọn ít nhất 1 kênh!")
                    with ra2:
                        st.button("✅ Chọn Tất Cả", key="btn_sel_all_t3_rej", on_click=cb_select_all, args=(display_rejected,), use_container_width=True)
                    with ra3:
                        st.button("❌ Bỏ Chọn", key="btn_clear_sel_t3_rej", on_click=cb_clear_all, use_container_width=True)

                items_per_page_rej = 20
                total_pages_rej = max(1, (len(display_rejected) + items_per_page_rej - 1) // items_per_page_rej)
                
                col_prj1, col_prj2 = st.columns([2, 8])
                with col_prj1:
                    page_rej = st.number_input("Trang (Kênh Bị Loại):", min_value=1, max_value=total_pages_rej, value=1, step=1, key="page_rej_t3")
                with col_prj2:
                    st.write("")
                    st.markdown(f"📄 **Trang {page_rej} / {total_pages_rej}** *(Hiển thị {min(20, len(display_rejected))} / {len(display_rejected)} kênh)*")

                start_idx_rej = (page_rej - 1) * items_per_page_rej
                paged_rejected = display_rejected[start_idx_rej:start_idx_rej + items_per_page_rej]

                for idx, item in enumerate(paged_rejected):
                    p_id = to_pure_id(item["Handle"])
                    is_active = (p_id == st.session_state.get('active_inspected_handle'))
                    is_in_cart = p_id in st.session_state['cart']
                    stt_num_rej = start_idx_rej + idx + 1
                    
                    with st.container(border=True):
                        if is_active:
                            st.markdown('<div class="active-card-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="active-banner-tag">🔍 ĐANG XEM 6 VIDEO MỚI CỦA KÊNH NÀY</div>', unsafe_allow_html=True)
                        elif is_in_cart:
                            st.markdown('<div class="in-cart-marker"></div>', unsafe_allow_html=True)
                            st.markdown('<div class="in-cart-banner-tag">🛒 ĐÃ CÓ TRONG GIỎ HÀNG</div>', unsafe_allow_html=True)

                        c0, c1, c2, c3, c4 = st.columns([0.4, 2.2, 3.0, 1.8, 3.0])
                        with c0:
                            st.checkbox("", key=f"chk_r_{p_id}_{st.session_state['chk_counter']}", value=(p_id in st.session_state['selected_channels']), on_change=toggle_select_channel, args=(p_id,))
                        with c1:
                            st.markdown(f"<h3 style='margin:0; font-weight:800; font-size:1.3rem;'><span class='badge-stt'>#{stt_num_rej}</span><a href='{item.get('Link Kênh', f'https://youtube.com/@{p_id}')}' style='text-decoration:none;'>{item['Handle']}</a></h3>", unsafe_allow_html=True)
                            st.write(f"**{item.get('Tên Kênh', 'N/A')}**")
                            if st.button("👁️ Xem 6 Video Mới", key=f"btn_prev_rej_{p_id}", on_click=set_active_inspected_channel, args=(p_id,)):
                                show_video_dialog(p_id, pre_fetched_videos=item.get('recent_videos'))
                        with c2:
                            st.write(f"👥 **Subs:** `{item.get('Subscribers', 'N/A')}` | 🌍 **Q.Gia:** `{item.get('Quốc gia', '')}`")
                            st.write(f"🎬 **Tổng Video:** `{item.get('Tổng Số Video', '')}` | 📅 **Mới nhất:** `{item.get('Video Gần Nhất', '')}`")
                            st.markdown(render_social_badges_html(item.get("Socials", {})), unsafe_allow_html=True)
                        with c3:
                            st.write(f"**Database:** {item.get('Trạng Thái DB', '')}")
                            st.markdown(f"❌ **Lý do:** <span style='color:#D95F26; font-weight:700;'>{item.get('Lý do loại', '')}</span>", unsafe_allow_html=True)
                        with c4:
                            bc1, bc2 = st.columns(2)
                            if is_in_cart:
                                current_tag = st.session_state['cart'][p_id].get("Tag", "📌 Chưa phân loại")
                                new_tag = st.selectbox("Gắn Nhãn:", ["📌 Chưa phân loại", "🔥 Ưu tiên làm", "📩 Đã liên hệ", "⏳ Đang chờ duyệt", "✅ Đã chốt", "❌ Bỏ qua"], index=["📌 Chưa phân loại", "🔥 Ưu tiên làm", "📩 Đã liên hệ", "⏳ Đang chờ duyệt", "✅ Đã chốt", "❌ Bỏ qua"].index(current_tag), key=f"tag_r_{p_id}")
                                if new_tag != current_tag:
                                    st.session_state['cart'][p_id]["Tag"] = new_tag
                                    add_to_cart_db(p_id, st.session_state['cart'][p_id])
                                if bc1.button("❌ Bỏ Giỏ", key=f"rm_r_{p_id}", use_container_width=True):
                                    remove_from_cart_db(p_id)
                                    del st.session_state['cart'][p_id]
                                    st.rerun()
                            else:
                                if bc1.button("🛒 Thêm Giỏ", key=f"add_r_{p_id}", use_container_width=True):
                                    item_data = dict(item); item_data["Tag"] = "📌 Chưa phân loại"
                                    st.session_state['cart'][p_id] = item_data
                                    add_to_cart_db(p_id, item_data)
                                    st.rerun()

                            audit_key = f"audit_file_{p_id}"
                            if audit_key in st.session_state:
                                bc2.download_button("📥 Tải Audit", data=st.session_state[audit_key]["bytes"], file_name=st.session_state[audit_key]["filename"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_r_{p_id}", use_container_width=True)
                            else:
                                if bc2.button("📄 Tạo Audit", key=f"btn_r_{p_id}", use_container_width=True):
                                    with st.spinner("Đang dựng Audit..."):
                                        b_data, f_name = run_single_channel_audit(p_id)
                                        if b_data:
                                            supabase.table("channels").upsert([{"handle": p_id, "youtuber_name": item.get('Tên Kênh', p_id.upper()), "source": "Smart Finder Audit"}], on_conflict="handle").execute()
                                            st.session_state['audit_success_msg'] = f"🎉 Đã lưu **@{p_id}** vào Database!"
                                            st.session_state[audit_key] = {"bytes": b_data, "filename": f_name}
                                            st.rerun()

                            bc3, bc4 = st.columns(2)
                            if bc3.button("🎯 Đào Sâu", key=f"deep_r_{p_id}", type="secondary", use_container_width=True):
                                cid_deep = get_channel_id_by_handle(p_id)
                                if cid_deep:
                                    ext_deep = extract_channel_master_keywords(cid_deep)
                                    st.session_state['pending_keywords'] = ", ".join(ext_deep['master_keywords'][:6])
                                    st.session_state['pending_seed_input'] = f"@{p_id}"
                                    st.session_state['trigger_deep_search_now'] = True
                                    st.rerun()

                            if bc4.button("🗑️ Xóa", key=f"del_r_{p_id}", use_container_width=True):
                                delete_channel_from_system(p_id)
                                st.toast(f"🗑️ Đã xóa kênh @{p_id}!")
                                st.rerun()

    render_shared_cart_ui(key_suffix="tab3")

# --- TAB 4: UPLOAD & UPDATE DATABASE DIRECTLY FROM EXCEL / ZIP / TXT ---
with tab4:
    st.markdown("<h3 style='font-weight: 700; margin-top: 15px;'>📤 Upload file .ZIP, .TXT hoặc .XLSX để cập nhật Database</h3>", unsafe_allow_html=True)
    st.caption("💡 *Hỗ trợ tải lên trực tiếp các file Excel báo cáo lẻ (.xlsx), file nén .ZIP hoặc file danh sách .TXT.*")
    
    uploaded_files = st.file_uploader("Kéo thả file `.zip` (chứa các báo cáo), file `.txt` hoặc file Excel báo cáo `.xlsx` vào đây:", type=["zip", "txt", "xlsx", "xls"], accept_multiple_files=True)
    
    if uploaded_files and st.button("🚀 Bắt đầu xử lý & Nạp vào Database", type="primary"):
        new_handles_to_insert = []
        skipped_details = []
        
        for file in uploaded_files:
            file_name = file.name.lower()
            if file_name.endswith('.zip'):
                with zipfile.ZipFile(file, 'r') as zip_ref:
                    extract_path = "temp_zip_extract"
                    zip_ref.extractall(extract_path)
                    for root, _, filenames in os.walk(extract_path):
                        for fn in filenames:
                            if fn.startswith('~$') or fn.startswith('._'):
                                skipped_details.append({"File": fn, "Lý do": "File ẩn / tạm của hệ điều hành"})
                                continue
                            if fn.endswith('.xlsx') or fn.endswith('.xls'):
                                h = extract_handle_from_filename(fn)
                                if h: new_handles_to_insert.append({"handle": h, "youtuber_name": h.upper(), "source": file.name, "filename": fn})
                                else: skipped_details.append({"File": fn, "Lý do": "Không trích xuất được Handle từ tên file"})
            elif file_name.endswith('.txt'):
                content = file.read().decode("utf-8", errors="ignore")
                for line in content.splitlines():
                    h = to_pure_id(line)
                    if h: new_handles_to_insert.append({"handle": h, "youtuber_name": h.upper(), "source": file.name, "filename": file.name})
            elif file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                h_from_fn = extract_handle_from_filename(file.name)
                
                file.seek(0)
                is_report = False
                try:
                    df_head = pd.read_excel(file, nrows=2, header=None)
                    if not df_head.empty:
                        first_cell = str(df_head.iloc[0, 0]).upper()
                        if "YOUTUBE CHANNEL SUMMARY REPORT" in first_cell or h_from_fn:
                            is_report = True
                            if not h_from_fn:
                                m = re.search(r'([^\s]+)\s+YOUTUBE CHANNEL SUMMARY REPORT', first_cell, re.IGNORECASE)
                                if m: h_from_fn = to_pure_id(m.group(1))
                except Exception: pass
                
                if is_report and h_from_fn:
                    new_handles_to_insert.append({"handle": h_from_fn, "youtuber_name": h_from_fn.upper(), "source": file.name, "filename": file.name})
                else:
                    file.seek(0)
                    try:
                        df_excel = pd.read_excel(file)
                        for col in df_excel.columns:
                            for val in df_excel[col].dropna():
                                p = to_pure_id(val)
                                if p:
                                    new_handles_to_insert.append({"handle": p, "youtuber_name": p.upper(), "source": file.name, "filename": file.name})
                    except Exception: pass

        if new_handles_to_insert:
            df_raw = pd.DataFrame(new_handles_to_insert)
            
            duplicated_rows = df_raw[df_raw.duplicated(subset=["handle"], keep="first")]
            for _, d_row in duplicated_rows.iterrows():
                skipped_details.append({"File": d_row.get("filename", d_row["source"]), "Handle": f"@{d_row['handle']}", "Lý do": "Trùng lặp Handle với 1 file khác trong gói ZIP"})

            df_insert = df_raw.drop_duplicates(subset=["handle"])
            supabase.table("channels").upsert(df_insert[["handle", "youtuber_name", "source"]].to_dict(orient="records"), on_conflict="handle").execute()
            
            st.success(f"🎉 Đã xử lý & đồng bộ thành công {len(df_insert)} Handle duy nhất vào Database đám mây Supabase!")
            
            if skipped_details:
                with st.expander(f"⚠️ Chi tiết ({len(skipped_details)}) file / Handle trùng lặp bị bỏ qua"):
                    st.dataframe(pd.DataFrame(skipped_details), use_container_width=True)
        else:
            st.warning("⚠️ Không tìm thấy Handle hợp lệ nào trong các file đã tải lên!")

# --- MULTI-THREADED CRM DATABASE VIEWER WITH SAFE INSTANT ALL-SUBS & CACHING ---
with tab5:
    st.markdown("<h3 style='font-weight: 700; margin-top: 15px;'>📊 Quản lý Database CRM Kênh</h3>", unsafe_allow_html=True)
    res = supabase.table("channels").select("*").execute()
    if res.data:
        df_all = pd.DataFrame(res.data)
        
        c_top1, c_top2 = st.columns([7, 3])
        with c_top1:
            st.markdown(f"Tổng số kênh hiện có trong DB: <span style='font-weight:800; color:#D95F26;'>{len(df_all)}</span>", unsafe_allow_html=True)
        with c_top2:
            if st.button("💣 Xóa Vĩnh Viễn Toàn Bộ DB", use_container_width=True, key="btn_trigger_wipe_db"):
                confirm_clear_db_dialog()

        st.divider()
        st.markdown("#### 🔍 Bộ Lọc Database Chuyên Sâu:")
        fc1, fc2, fc3, fc4 = st.columns([2, 2, 2, 2])
        with fc1:
            search_db = st.text_input("Tìm kiếm thủ công theo Handle/Tên:", "")
        with fc2:
            sub_range_options = ["-- Tất Cả Mốc Subs --", "< 100K Subs", "100K - 500K Subs", "500K - 1M Subs", "> 1M Subs"]
            sel_sub_range = st.selectbox("Lọc theo Mốc Subscribers:", options=sub_range_options)
        with fc3:
            source_options = ["-- Tất Cả Nguồn --"] + list(df_all['source'].dropna().unique()) if 'source' in df_all.columns else ["-- Tất Cả Nguồn --"]
            sel_source = st.selectbox("Lọc theo Nguồn Nạp (File ZIP / Source):", options=source_options)
        with fc4:
            view_mode = st.radio("Chế độ hiển thị:", ["🎨 Card View (Thẻ chi tiết)", "📊 Table Grid View (Bảng nén gọn)"], horizontal=True)

        df_pre = df_all.copy()
        if search_db:
            df_pre = df_pre[
                df_pre['handle'].str.contains(search_db, case=False, na=False) | 
                df_pre['youtuber_name'].str.contains(search_db, case=False, na=False)
            ]
        if sel_source != "-- Tất Cả Nguồn --":
            df_pre = df_pre[df_pre['source'] == sel_source]

        crm_meta_map = {}

        if sel_sub_range != "-- Tất Cả Mốc Subs --":
            cache_key = f"crm_cache_{search_db}_{sel_sub_range}_{sel_source}"
            if cache_key in st.session_state:
                df_filtered, crm_meta_map = st.session_state[cache_key]
            else:
                handles_to_check = [to_pure_id(h) for h in df_pre['handle'].tolist() if to_pure_id(h)]
                tot_h = len(handles_to_check)
                matched_handles = []
                
                if tot_h > 0:
                    prog_bar_db = st.progress(0)
                    stat_txt_db = st.empty()
                    comp_h = 0

                    with ThreadPoolExecutor(max_workers=10) as executor:
                        futures = [executor.submit(process_single_crm_channel_meta, p_h) for p_h in handles_to_check]
                        for future in as_completed(futures):
                            p_h, meta = future.result()
                            crm_meta_map[p_h] = meta
                            s_num = meta['sub_count']
                            
                            is_match = False
                            if sel_sub_range == "< 100K Subs" and (0 <= s_num < 100000): is_match = True
                            elif sel_sub_range == "100K - 500K Subs" and (100000 <= s_num < 500000): is_match = True
                            elif sel_sub_range == "500K - 1M Subs" and (500000 <= s_num < 1000000): is_match = True
                            elif sel_sub_range == "> 1M Subs" and s_num >= 1000000: is_match = True
                            
                            if is_match: matched_handles.append(p_h)
                            
                            comp_h += 1
                            prog_bar_db.progress(comp_h / tot_h)
                            stat_txt_db.markdown(f"⏳ **Đang phân tích siêu tốc:** `{comp_h}/{tot_h}` kênh | 🎯 **Khớp điều kiện:** `{len(matched_handles)}` kênh")

                    prog_bar_db.empty()
                    stat_txt_db.empty()

                    df_filtered = df_pre[df_pre['handle'].apply(to_pure_id).isin(matched_handles)]
                else: df_filtered = df_pre

                st.session_state[cache_key] = (df_filtered, crm_meta_map)
        else:
            df_filtered = df_pre

        st.caption(f"🎯 Kết quả khớp: **{len(df_filtered)}** / {len(df_all)} kênh")

        # RENDER MODE 1: TABLE GRID VIEW
        if view_mode == "📊 Table Grid View (Bảng nén gọn)":
            df_grid = df_filtered.copy()
            df_grid['Link Kênh'] = df_grid['handle'].apply(lambda h: f"https://youtube.com/@{to_pure_id(h)}" if to_pure_id(h) else "")
            df_grid['Tab Videos'] = df_grid['handle'].apply(lambda h: f"https://youtube.com/@{to_pure_id(h)}/videos" if to_pure_id(h) else "")
            
            df_grid.index = range(1, len(df_grid) + 1)
            st.dataframe(df_grid[['handle', 'youtuber_name', 'source', 'Link Kênh', 'Tab Videos']], use_container_width=True, column_config={
                "handle": st.column_config.TextColumn("Handle"),
                "youtuber_name": st.column_config.TextColumn("Tên YouTuber"),
                "source": st.column_config.TextColumn("Nguồn Dữ Liệu"),
                "Link Kênh": st.column_config.LinkColumn("Trang Chủ", display_text="🏠 Kênh"),
                "Tab Videos": st.column_config.LinkColumn("Tab Videos", display_text="🎬 Videos")
            })
        else:
            # RENDER MODE 2: CARD VIEW WITH PAGINATION & PAGE COUNTER
            items_per_page = 20
            total_pages = max(1, (len(df_filtered) + items_per_page - 1) // items_per_page)
            
            col_db_p1, col_db_p2 = st.columns([2, 8])
            with col_db_p1:
                page = st.number_input("Trang:", min_value=1, max_value=total_pages, value=1, step=1, key="page_db_viewer")
            with col_db_p2:
                st.write("")
                st.markdown(f"📄 **Trang {int(page)} / {total_pages}** *(Hiển thị {min(20, len(df_filtered))} / {len(df_filtered)} kênh)*")

            start_idx = (int(page) - 1) * items_per_page
            end_idx = start_idx + items_per_page
            page_data = df_filtered.iloc[start_idx:end_idx]

            # SAFE STICKY FLOATING ACTION BAR FOR DB
            with st.container(border=True):
                st.markdown('<div class="sticky-action-bar"></div>', unsafe_allow_html=True)
                cnt_total_sel_db = len(st.session_state['selected_channels'])
                db_act1, db_act2, db_act3, db_act4 = st.columns([3, 2, 2, 1.5])
                with db_act1:
                    if st.button(f"🗑️ Xóa ({cnt_total_sel_db}) Kênh Đã Chọn Khỏi DB", key="btn_del_sel_db", use_container_width=True):
                        if cnt_total_sel_db > 0:
                            for p_id in list(st.session_state['selected_channels']):
                                delete_channel_from_system(p_id)
                            cb_clear_all()
                            st.success(f"🗑️ Đã xóa {cnt_total_sel_db} kênh khỏi Database!")
                            st.rerun()
                        else: st.warning("Vui lòng tick chọn ít nhất 1 kênh trong DB!")
                with db_act2:
                    st.button("✅ Chọn Tất Cả (Trang Này)", key="btn_sel_page_db", on_click=cb_select_all, args=(page_data.to_dict('records'),), use_container_width=True)
                with db_act3:
                    st.button("✅ Chọn Tất Cả (Toàn DB)", key="btn_sel_all_db", on_click=cb_select_all, args=(df_filtered.to_dict('records'),), use_container_width=True)
                with db_act4:
                    st.button("❌ Bỏ Chọn", key="btn_clear_sel_db", on_click=cb_clear_all, use_container_width=True)

            st.divider()
            for idx, row in page_data.iterrows():
                p_id = to_pure_id(row['handle'])
                is_active = (p_id == st.session_state.get('active_inspected_handle'))
                stt_num_db = start_idx + idx + 1
                
                crm_meta = crm_meta_map.get(p_id)
                if not crm_meta or crm_meta.get("sub_count", -1) == -1:
                    crm_meta = get_channel_crm_meta(p_id)
                
                with st.container(border=True):
                    if is_active:
                        st.markdown('<div class="active-card-marker"></div>', unsafe_allow_html=True)
                        st.markdown('<div class="active-banner-tag">🔍 ĐANG XEM 6 VIDEO MỚI CỦA KÊNH NÀY</div>', unsafe_allow_html=True)

                    c0, c1, c2, c3 = st.columns([0.4, 3.8, 3.8, 2.0])
                    with c0:
                        st.checkbox("", key=f"chk_db_{p_id}_{st.session_state['chk_counter']}", value=(p_id in st.session_state['selected_channels']), on_change=toggle_select_channel, args=(p_id,))
                    with c1:
                        st.markdown(f"<h3 style='margin:0; font-weight:800; font-size:1.3rem;'><span class='badge-stt'>#{stt_num_db}</span><a href='https://youtube.com/@{p_id}' style='text-decoration:none;'>@{p_id}</a></h3>", unsafe_allow_html=True)
                        st.write(f"**Tên YouTuber:** {row.get('youtuber_name', 'N/A')}")
                        if st.button("👁️ Xem 6 Video Mới", key=f"btn_prev_db_{idx}_{p_id}", on_click=set_active_inspected_channel, args=(p_id,)):
                            show_video_dialog(p_id)
                    with c2:
                        st.write(f"👥 **Subs:** `{crm_meta['sub_str']}` | 🌍 **Q.Gia:** `{crm_meta['country']}`")
                        st.write(f"📁 **Nguồn dữ liệu:** {row.get('source', 'N/A')}")
                        st.markdown(render_social_badges_html(crm_meta.get("socials", {})), unsafe_allow_html=True)
                    with c3:
                        if st.button("🗑️ Xóa DB", key=f"del_db_{idx}_{p_id}", use_container_width=True):
                            delete_channel_from_system(p_id)
                            st.toast(f"🗑️ Đã xóa kênh @{p_id} khỏi Database!")
                            st.rerun()

        st.divider()
        csv = df_all.to_csv(index=False).encode('utf-8')
        st.download_button("📥 Tải về toàn bộ Database (CSV)", data=csv, file_name="master_youtube_database.csv", mime="text/csv", type="primary")
    else:
        st.info("Database hiện đang trống!")

with tab6:
    st.markdown("<h3 style='font-weight: 700; margin-top: 15px;'>✨ Soi Từ Khóa Kênh (Channel & Video Tags SEO Inspector)</h3>", unsafe_allow_html=True)
    inspect_handle_input = st.text_input("Nhập Handle Kênh cần soi:", value="@NickDiGiovanni")
    if inspect_handle_input and st.button("🔍 Soi Từ Khóa Ngay", type="primary"):
        pure_inspect = to_pure_id(inspect_handle_input)
        if pure_inspect:
            try:
                cid_insp = get_channel_id_by_handle(pure_inspect)
                if not cid_insp: st.error("Không tìm thấy Channel ID cho kênh này!")
                else:
                    ext_data = extract_channel_master_keywords(cid_insp)
                    st.session_state['pending_keywords'] = ", ".join(ext_data['master_keywords'])
                    st.session_state['last_inspected_data'] = ext_data
                    st.session_state['last_inspected_handle'] = pure_inspect
                    st.rerun()
            except Exception as e: st.error(f"Lỗi khi soi từ khóa: {e}")

    if 'last_inspected_data' in st.session_state:
        ext_data, pure_inspect = st.session_state['last_inspected_data'], st.session_state.get('last_inspected_handle', '')
        st.divider(); st.success(f"✨ Đã liên kết tự động bộ từ khóa này sang Tab 3 ('Săn Kênh Tương Tự')!")
        st.markdown(f"### 🏷️ Dữ Liệu Từ Khóa Của Kênh `@{pure_inspect}`")
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            st.markdown("#### 🔑 Thẻ Từ Khóa Ẩn (Channel Keywords):")
            for kw in ext_data['channel_keywords']: st.write(f"• `{kw}`")
        with col_t2:
            st.markdown("#### 📌 Top Video Tags Xuất Hiện Nhiều Nhất:")
            for tag in ext_data['top_tags']: st.write(f"• `{tag}`")
        st.code(", ".join(ext_data['master_keywords']), language="text")
